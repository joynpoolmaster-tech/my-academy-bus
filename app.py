# app.py (지점별 권한 분리 완전 적용 버전)
# 주요 수정사항:
# 1. 지점별 데이터 필터링 (마스터는 전체, 관리자는 자신의 지점만)
# 2. 권한 체크 강화 (모든 CRUD 작업에 권한 검증)
# 3. 차량-지점 연결 기능 추가
# 4. 기사 페이지 지점별 필터링

# app.py 파일 맨 위에 추가
import json
from flask import Flask, request, render_template, redirect, url_for, flash, jsonify, send_file, session
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import os
from database import db
import os                    # ← 이거 추가
from config import Config    # ← 이거 추가
from collections import defaultdict
import pandas as pd
import io
from sqlalchemy import func
from functools import wraps

app = Flask(__name__)
app.config.from_object(Config)    # ← 이거 추가
basedir = os.path.abspath(os.path.dirname(__file__))
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(basedir, 'database.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'fallback-secret-key-for-development')
db.init_app(app)

from models import User, Student, Class, TimeSlot, Vehicle, DispatchResult, Branch
# --- 로그인 확인 데코레이터 ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash("로그인이 필요합니다.", "warning")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def master_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') != 'master':
            flash("마스터 관리자만 접근할 수 있는 페이지입니다.", "danger")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function



def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session or session.get('role') not in ['master', 'admin']:
            flash("관리자만 접근할 수 있는 페이지입니다.", "danger")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# 🔹 여기에 새로운 권한 체크 함수들 추가
def check_user_permission_for_student(current_user, student):
    """사용자가 해당 학생에 대한 권한이 있는지 확인"""
    if current_user.role == 'master':
        return True
    elif current_user.role == 'admin':
        # branch_id 우선, branch_name 백업
        if student.branch_id == current_user.branch_id:
            return True
        elif (hasattr(current_user, 'managed_branch') and 
              current_user.managed_branch and 
              student.branch_name == current_user.managed_branch.name):
            return True
    elif current_user.role == 'driver':
        # 기사는 자신의 차량에 배정된 학생만
        if hasattr(current_user, 'vehicle') and current_user.vehicle:
            return student.branch_id == current_user.vehicle.branch_id
    return False

def check_user_permission_for_vehicle(current_user, vehicle):
    """사용자가 해당 차량에 대한 권한이 있는지 확인"""
    if current_user.role == 'master':
        return True
    elif current_user.role == 'admin':
        return vehicle.branch_id == current_user.branch_id
    elif current_user.role == 'driver':
        # 기사는 자신의 차량만
        return hasattr(current_user, 'vehicle') and current_user.vehicle.id == vehicle.id
    return False

def check_user_permission_for_class(current_user, class_item):
    """사용자가 해당 클래스에 대한 권한이 있는지 확인"""
    if current_user.role == 'master':
        return True
    elif current_user.role == 'admin':
        return class_item.branch_id == current_user.branch_id
    return False

def check_user_permission_for_branch(current_user, branch_id):
    """사용자가 해당 지점에 대한 권한이 있는지 확인"""
    if current_user.role == 'master':
        return True
    elif current_user.role == 'admin':
        return current_user.branch_id == int(branch_id)
    return False
def setup_initial_accounts():
    with app.app_context():
        try:
            if not User.query.filter_by(email='master@joypool.com').first():
                master_admin = User(email='master@joypool.com', name='마스터 관리자', role='master')
                master_admin.set_password(os.getenv('MASTER_PASSWORD', 'temp-password-change-immediately'))
                db.session.add(master_admin)
            if not User.query.filter_by(email='driver@joypool.com').first():
                test_driver = User(email='driver@joypool.com', name='테스트 기사', phone='010-1234-5678', role='driver')
                test_driver.set_password(os.getenv('DRIVER_PASSWORD', 'temp-password-change-immediately'))
                db.session.add(test_driver)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            print(f"초기 계정 설정 오류: {e}")

@app.route('/')
def index():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        if not email or not password:
            flash("이메일과 비밀번호를 모두 입력해주세요.", "danger")
            return redirect(url_for('login'))
            
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(password):
            session['user_id'] = user.id
            session['user_name'] = user.name
            session['role'] = user.role
            
            flash(f"{user.name}님, 환영합니다!", "success")
            if user.role == 'master':
                return redirect(url_for('manage_branches'))
            elif user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            elif user.role == 'driver':
                return redirect(url_for('driver_view_route'))
            else:  # student
                flash("학생 전용 페이지는 아직 구현되지 않았습니다.", "info")
                return redirect(url_for('login'))
        else:
            flash("이메일 또는 비밀번호가 올바르지 않습니다.", "danger")
            return redirect(url_for('login'))
    return render_template('auth/login.html')

@app.route('/logout')
@login_required
def logout():
    session.clear()
    # 🔹 수정: 로그아웃 메시지를 session.clear() 후에 추가
    flash("성공적으로 로그아웃되었습니다.", "info")
    return redirect(url_for('login'))

@app.route('/signup', methods=['GET', 'POST'])
def signup_member():
    if request.method == 'POST':
        try:
            form = request.form
            name = form.get('name')
            email = form.get('email')
            password = form.get('password')
            confirm_password = form.get('confirm_password')
            phone = form.get('phone')
            branch_id = form.get('branch_id')
            class_id = form.get('class_id')
            
            # 유효성 검사
            if not all([name, email, password, confirm_password, phone, branch_id, class_id]):
                flash("모든 필수 필드를 입력해주세요.", "danger")
                return redirect(url_for('signup_member'))
            
            if password != confirm_password:
                flash("비밀번호가 일치하지 않습니다.", "danger")
                return redirect(url_for('signup_member'))
                
            if User.query.filter_by(email=email).first():
                flash("이미 가입된 이메일 주소입니다.", "danger")
                return redirect(url_for('signup_member'))

            # 수강 기간 계산
            start_date_str = form.get('start_date')
            duration_months = int(form.get('duration', 1))
            end_date = None
            start_date_obj = None
            
            if start_date_str:
                start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = start_date_obj + relativedelta(months=duration_months)

            # 사용자 생성
            new_user = User(email=email, name=name, phone=phone, role='student')
            new_user.set_password(password)
            db.session.add(new_user)
            db.session.flush()  # ID 생성을 위해 flush

            # 클래스 및 지점 정보 조회
            class_item = Class.query.get(class_id)
            branch_item = Branch.query.get(branch_id)
            
            if not class_item or not branch_item:
                raise Exception("선택한 클래스 또는 지점이 존재하지 않습니다.")
            
            # 🔹 수정: branch_id도 함께 저장
            new_student_info = Student(
                user_id=new_user.id,
                branch_id=int(branch_id),  # 🔹 추가
                branch_name=branch_item.name,
                class_name=class_item.name,
                time_slot=form.get('time_slot'),
                address=form.get('address'),
                status='pending',
                start_date=start_date_obj,
                emergency_contact=form.get('emergency_contact'),
                end_date=end_date
            )
            db.session.add(new_student_info)
            db.session.commit()
            
            flash("회원가입이 성공적으로 완료되었습니다. 관리자 승인 후 수강이 가능합니다.", "success")
            return redirect(url_for('login'))
            
        except Exception as e:
            db.session.rollback()
            flash(f"회원가입 중 오류가 발생했습니다: {str(e)}", "danger")
            return redirect(url_for('signup_member'))
            
    all_branches = Branch.query.all()
    return render_template("auth/signup_member.html", branches=all_branches)

# ----------------------------------------------------
# 🔹 마스터 관리자 전용 라우트
# ----------------------------------------------------
@app.route('/master/branches')
@master_required
def manage_branches():
    all_branches = Branch.query.all()
    return render_template('master/manage_branches.html', branches=all_branches)

@app.route('/master/add_branch', methods=['POST'])
@master_required
def add_branch():
    try:
        name = request.form.get('name')
        password = request.form.get('password')
        
        if not name or not password:
            flash("지점 이름과 비밀번호를 모두 입력해주세요.", "danger")
            return redirect(url_for('manage_branches'))
            
        master_user = User.query.get(session['user_id'])
        if not master_user or not master_user.check_password(password):
            flash("마스터 비밀번호가 일치하지 않습니다.", "danger")
            return redirect(url_for('manage_branches'))

        if Branch.query.filter_by(name=name).first():
            flash("이미 존재하는 지점명입니다.", "danger")
            return redirect(url_for('manage_branches'))
            
        new_branch = Branch(name=name)
        db.session.add(new_branch)
        db.session.commit()
        flash(f"'{name}' 지점이 생성되었습니다.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"지점 생성 중 오류가 발생했습니다: {str(e)}", "danger")
        
    return redirect(url_for('manage_branches'))

# 기존 delete_branch 함수를 이것으로 교체
@app.route('/master/delete_branch/<int:branch_id>', methods=['POST'])
@master_required
def delete_branch(branch_id):
    try:
        password = request.form.get('password')
        if not password:
            flash("비밀번호를 입력해주세요.", "danger")
            return redirect(url_for('manage_branches'))
            
        master_user = User.query.get(session['user_id'])
        if not master_user.check_password(password):
            flash("마스터 비밀번호가 일치하지 않습니다.", "danger")
            return redirect(url_for('manage_branches'))
            
        branch = Branch.query.get_or_404(branch_id)
        
        # 🔹 개선: 삭제 가능 여부 상세 체크
        # 1. 등록된 학생 확인
        students_count = Student.query.filter_by(branch_id=branch_id).count()
        if students_count > 0:
            flash(f"'{branch.name}' 지점에 등록된 학생({students_count}명)이 있어 삭제할 수 없습니다. 먼저 학생들을 다른 지점으로 이전하거나 삭제해주세요.", "danger")
            return redirect(url_for('manage_branches'))
        
        # 2. 차량 확인
        vehicles_count = Vehicle.query.filter_by(branch_id=branch_id).count()
        if vehicles_count > 0:
            flash(f"'{branch.name}' 지점에 등록된 차량({vehicles_count}대)이 있어 삭제할 수 없습니다. 먼저 차량을 삭제해주세요.", "danger")
            return redirect(url_for('manage_branches'))
        
        # 3. 배차 기록 확인
        branch_vehicle_ids = [v.id for v in Vehicle.query.filter_by(branch_id=branch_id).all()]
        dispatch_count = DispatchResult.query.filter(DispatchResult.vehicle_id.in_(branch_vehicle_ids)).count()
        if dispatch_count > 0:
            flash(f"'{branch.name}' 지점에 배차 기록이 있어 삭제할 수 없습니다.", "danger")
            return redirect(url_for('manage_branches'))
        
        branch_name = branch.name
        
        # 4. 관련 데이터 삭제 (순서 중요)
        # 클래스와 시간대 삭제
        related_classes = Class.query.filter_by(branch_id=branch_id).all()
        for cls in related_classes:
            TimeSlot.query.filter_by(class_id=cls.id).delete()
            db.session.delete(cls)
        
        # 관리자 계정 삭제
        related_admins = User.query.filter_by(branch_id=branch_id, role='admin').all()
        for admin in related_admins:
            db.session.delete(admin)
        
        # 기사 계정 처리 (driver_branch_id 초기화)
        related_drivers = User.query.filter_by(driver_branch_id=branch_id, role='driver').all()
        for driver in related_drivers:
            driver.driver_branch_id = None
        
        # 지점 삭제
        db.session.delete(branch)
        db.session.commit()
        flash(f"'{branch_name}' 지점과 관련된 모든 데이터가 삭제되었습니다.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"지점 삭제 중 오류가 발생했습니다: {str(e)}", "danger")
        
    return redirect(url_for('manage_branches'))

# 🔹 새로 추가: 관리자 삭제 기능
@app.route('/master/delete_admin/<int:admin_id>', methods=['POST'])
@master_required
def delete_admin(admin_id):
    try:
        password = request.form.get('password')
        if not password:
            flash("비밀번호를 입력해주세요.", "danger")
            return redirect(url_for('manage_branches'))
            
        master_user = User.query.get(session['user_id'])
        if not master_user.check_password(password):
            flash("마스터 비밀번호가 일치하지 않습니다.", "danger")
            return redirect(url_for('manage_branches'))
            
        admin_to_delete = User.query.get_or_404(admin_id)
        
        # 마스터는 삭제할 수 없음
        if admin_to_delete.role == 'master':
            flash("마스터 계정은 삭제할 수 없습니다.", "danger")
            return redirect(url_for('manage_branches'))
        
        admin_name = admin_to_delete.name
        branch_name = admin_to_delete.managed_branch.name if admin_to_delete.managed_branch else "알 수 없음"
        
        db.session.delete(admin_to_delete)
        db.session.commit()
        flash(f"'{branch_name}' 지점의 관리자 '{admin_name}' 계정이 삭제되었습니다.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"관리자 삭제 중 오류가 발생했습니다: {str(e)}", "danger")
        
    return redirect(url_for('manage_branches'))

# 🔹 새로 추가: 지점 상세 정보 API
@app.route('/api/branch_info/<int:branch_id>')
@master_required
def get_branch_info(branch_id):
    """지점 상세 정보 조회 (삭제 가능 여부 확인용)"""
    try:
        branch = Branch.query.get_or_404(branch_id)
        
        # 관련 데이터 개수 조회
        students_count = Student.query.filter_by(branch_id=branch_id).count()
        vehicles_count = Vehicle.query.filter_by(branch_id=branch_id).count()
        classes_count = Class.query.filter_by(branch_id=branch_id).count()
        admins_count = User.query.filter_by(branch_id=branch_id, role='admin').count()
        drivers_count = User.query.filter_by(driver_branch_id=branch_id, role='driver').count()
        
        # 배차 기록 확인
        branch_vehicle_ids = [v.id for v in Vehicle.query.filter_by(branch_id=branch_id).all()]
        dispatch_count = DispatchResult.query.filter(DispatchResult.vehicle_id.in_(branch_vehicle_ids)).count()
        
        # 삭제 가능 여부 판단
        can_delete = (students_count == 0 and vehicles_count == 0 and dispatch_count == 0)
        
        delete_warnings = []
        if students_count > 0:
            delete_warnings.append(f"등록된 학생 {students_count}명")
        if vehicles_count > 0:
            delete_warnings.append(f"등록된 차량 {vehicles_count}대")
        if dispatch_count > 0:
            delete_warnings.append(f"배차 기록 {dispatch_count}건")
        
        return jsonify({
            'branch_name': branch.name,
            'can_delete': can_delete,
            'students_count': students_count,
            'vehicles_count': vehicles_count,
            'classes_count': classes_count,
            'admins_count': admins_count,
            'drivers_count': drivers_count,
            'dispatch_count': dispatch_count,
            'warnings': delete_warnings
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/master/add_admin', methods=['POST'])
@master_required
def add_admin():
    try:
        form = request.form
        email = form.get('email')
        name = form.get('name')
        password = form.get('password')
        branch_id = form.get('branch_id')
        
        if not all([email, name, password, branch_id]):
            flash("모든 필드를 입력해주세요.", "danger")
            return redirect(url_for('manage_branches'))
            
        if User.query.filter_by(email=email).first():
            flash(f"이미 사용 중인 이메일({email})입니다.", "danger")
            return redirect(url_for('manage_branches'))
        
        # 지점 존재 확인
        branch = Branch.query.get(branch_id)
        if not branch:
            flash("선택한 지점이 존재하지 않습니다.", "danger")
            return redirect(url_for('manage_branches'))
        
        new_admin = User(
            email=email,
            name=name,
            role='admin',
            branch_id=int(branch_id)
        )
        new_admin.set_password(password)
        db.session.add(new_admin)
        db.session.commit()
        flash(f"관리자 '{new_admin.name}' 계정이 생성되었습니다.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"관리자 생성 중 오류가 발생했습니다: {str(e)}", "danger")
        
    return redirect(url_for('manage_branches'))

# ----------------------------------------------------
# 🔹 엑셀 관련 기능 라우트
# ----------------------------------------------------
@app.route('/admin/download_template')
@admin_required
def download_dynamic_template():
    try:
        current_user = User.query.get(session['user_id'])
        
        # 지점별 클래스 조회
        if current_user.role == 'master':
            classes = Class.query.all()
            branch_name = "전체지점"
        else:
            classes = Class.query.filter_by(branch_id=current_user.branch_id).all()
            branch_name = current_user.managed_branch.name if current_user.managed_branch else "미설정"
        
        # 새 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "회원명부"
        
        # 헤더 설정
        headers = ['번호', '이름', '이메일', '초기비밀번호', '연락처', '비상연락망', 
                  '주소', '지점명', '클래스명', '시간대', '수강시작일(YYYY-MM-DD)']
        ws.append(headers)
        
        # 클래스명 목록 생성 (시간대 포함)
        class_options = []
        for c in classes:
            for time_slot in c.time_slots:
                class_options.append(c.name)
        
        # 중복 제거
        unique_classes = list(set(class_options))
        
        if unique_classes:
            # 클래스명 드롭다운 설정
            dv_class = DataValidation(
                type="list", 
                formula1=f'"{",".join(unique_classes)}"',
                showErrorMessage=True,
                errorTitle="잘못된 클래스명",
                error="목록에서 선택해주세요"
            )
            ws.add_data_validation(dv_class)
            dv_class.add('I2:I1000')  # 클래스명 컬럼 (I열)
        
        # 지점명 드롭다운 설정
        dv_branch = DataValidation(
            type="list", 
            formula1=f'"{branch_name}"',
            showErrorMessage=True,
            errorTitle="잘못된 지점명", 
            error="해당 지점명만 사용 가능합니다"
        )
        ws.add_data_validation(dv_branch)
        dv_branch.add('H2:H1000')  # 지점명 컬럼 (H열)
        
        # 샘플 데이터 1줄 추가
        if unique_classes:
            sample_row = [
                1, "홍길동", "sample@example.com", 12345, 
                "010-1234-5678", "010-9876-5432", 
                f"경기도 용인시 기흥구 동천동 123-45", 
                branch_name, unique_classes[0], "07:00", "2025-08-07"
            ]
            ws.append(sample_row)
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 메모리에 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 파일명 생성
        filename = f"{branch_name}_회원명부_양식_{date.today().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output, 
            as_attachment=True, 
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f"양식 생성 중 오류가 발생했습니다: {str(e)}", "danger")
        return redirect(url_for('manage_students'))

# app.py의 upload_students 함수 개선 버전

@app.route('/admin/upload_students', methods=['POST'])
@admin_required
def upload_students():
    try:
        current_user = User.query.get(session['user_id'])
        file = request.files.get('student_file')
        
        print(f"🔹 업로드 시작 - 사용자: {current_user.name}, 지점: {current_user.branch_id}")
        
        if not file or file.filename == '':
            flash("파일이 선택되지 않았습니다.", "danger")
            return redirect(url_for('manage_students'))

        df = pd.read_excel(file)
        new_students_count = 0
        error_count = 0
        
        print(f"📊 엑셀 데이터: {len(df)}행")
        
        for index, row in df.iterrows():
            try:
                # 필수 데이터 확인
                if pd.isna(row['이메일']) or pd.isna(row['이름']):
                    print(f"❌ {index+1}행: 필수 데이터 누락")
                    error_count += 1
                    continue
                    
                if User.query.filter_by(email=row['이메일']).first():
                    print(f"⚠️ {index+1}행: 이미 존재하는 이메일 ({row['이메일']})")
                    continue

                # 사용자 생성
                new_user = User(
                    name=str(row['이름']),
                    email=str(row['이메일']),
                    phone=str(row['연락처']) if pd.notna(row['연락처']) else '',
                    role='student'
                )
                new_user.set_password(str(row['초기비밀번호']))
                db.session.add(new_user)
                db.session.flush()
                
                print(f"✅ 사용자 생성: {new_user.name} (ID: {new_user.id})")

                # 수강 시작일 처리
                start_date = None
                end_date = None
                if pd.notna(row.get('수강시작일(YYYY-MM-DD)')):
                    try:
                        start_date = pd.to_datetime(row['수강시작일(YYYY-MM-DD)']).date()
                        # 수강 기간이 있으면 종료일 계산
                        if pd.notna(row.get('수강기간(개월)')):
                            duration_months = int(row['수강기간(개월)'])
                            end_date = start_date + relativedelta(months=duration_months)
                        print(f"📅 기간: {start_date} ~ {end_date}")
                    except Exception as e:
                        print(f"⚠️ 날짜 변환 실패: {e}")
                
                # 🔹 핵심 개선: 지점명 매칭 로직 강화
                branch_name = str(row['지점명']).strip() if pd.notna(row['지점명']) else ''
                branch_id = None
                
                if branch_name:
                    # 1차: 정확한 매칭 시도
                    branch_item = Branch.query.filter_by(name=branch_name).first()
                    
                    if not branch_item:
                        # 2차: 부분 매칭 시도 (공백, 대소문자 무시)
                        all_branches = Branch.query.all()
                        for b in all_branches:
                            if branch_name.replace(' ', '').lower() in b.name.replace(' ', '').lower() or \
                               b.name.replace(' ', '').lower() in branch_name.replace(' ', '').lower():
                                branch_item = b
                                print(f"🔍 부분 매칭 성공: '{branch_name}' → '{b.name}'")
                                break
                    
                    if branch_item:
                        branch_id = branch_item.id
                        # 🔹 추가: 권한 체크 (일반 관리자는 자신의 지점만)
                        if current_user.role != 'master' and branch_id != current_user.branch_id:
                            print(f"❌ {index+1}행: 권한 없음 - 다른 지점 학생")
                            error_count += 1
                            continue
                    else:
                        # 3차: 관리자의 기본 지점 사용
                        if current_user.role != 'master' and current_user.branch_id:
                            branch_id = current_user.branch_id
                            managed_branch = Branch.query.get(current_user.branch_id)
                            branch_name = managed_branch.name if managed_branch else branch_name
                            print(f"🔄 기본 지점 사용: {branch_name} (ID: {branch_id})")
                        else:
                            print(f"❌ {index+1}행: 지점 '{branch_name}'을 찾을 수 없고 기본 지점도 없음")
                            error_count += 1
                            continue
                else:
                    # 지점명이 없으면 관리자의 기본 지점 사용
                    if current_user.role != 'master' and current_user.branch_id:
                        branch_id = current_user.branch_id
                        managed_branch = Branch.query.get(current_user.branch_id)
                        branch_name = managed_branch.name if managed_branch else f"지점{current_user.branch_id}"
                        print(f"🔄 기본 지점 사용: {branch_name} (ID: {branch_id})")
                    else:
                        print(f"❌ {index+1}행: 지점명이 없고 기본 지점도 설정되지 않음")
                        error_count += 1
                        continue
                
                print(f"🏢 최종 지점: '{branch_name}' (ID: {branch_id})")
                
                # 학생 정보 생성 (이제 반드시 실행됨)
                new_student = Student(
                    user_id=new_user.id,
                    branch_id=branch_id,
                    branch_name=branch_name,
                    class_name=str(row.get('클래스명', '')),
                    time_slot=str(row.get('시간대', '')),
                    address=str(row.get('주소', '')),
                    emergency_contact=str(row.get('비상연락망', '')),
                    start_date=start_date,
                    end_date=end_date,
                    status='approved'
                )
                db.session.add(new_student)
                print(f"✅ 학생 정보 생성: {new_user.name}")
                
                new_students_count += 1
                
            except Exception as e:
                print(f"❌ {index+1}행 처리 중 오류: {str(e)}")
                error_count += 1
                continue
        
        # 🔹 개선: 부분 성공도 커밋
        if new_students_count > 0:
            try:
                db.session.commit()
                print(f"💾 데이터베이스 커밋 완료")
            except Exception as e:
                print(f"❌ 커밋 실패: {e}")
                db.session.rollback()
                flash(f"데이터베이스 저장 중 오류가 발생했습니다: {e}", "danger")
                return redirect(url_for('manage_students'))
        
        # 결과 확인 및 메시지
        total_students = Student.query.count()
        if current_user.role == 'master':
            branch_students = total_students
        else:
            branch_students = Student.query.filter_by(branch_id=current_user.branch_id).count()
        
        print(f"📊 등록 후 현황: 전체 {total_students}명, 내 지점 {branch_students}명")
        
        if new_students_count > 0:
            message = f"✅ {new_students_count}명의 학생이 성공적으로 등록되었습니다!"
            if error_count > 0:
                message += f" ({error_count}건의 오류가 있었습니다.)"
            flash(message, "success")
        else:
            flash(f"등록된 학생이 없습니다. {error_count}건의 오류가 발생했습니다.", "warning")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ 전체 업로드 실패: {str(e)}")
        flash(f"파일 처리 중 치명적인 오류가 발생했습니다: {e}", "danger")
    
    return redirect(url_for('manage_students'))

@app.route('/admin/download_students')
@admin_required
def download_students():
    """지점별 등록된 회원 명부 다운로드"""
    try:
        current_user = User.query.get(session['user_id'])
        
        # 현재 지점의 학생들만 조회
        if current_user.role == 'master':
            # 마스터는 모든 학생
            students = Student.query.join(User).order_by(User.created_at.desc()).all()
        else:
            # 지점 관리자는 자기 지점만
            students = Student.query.filter_by(branch_id=current_user.branch_id).join(User).order_by(User.created_at.desc()).all()
        
        # 엑셀 데이터 생성
        data = []
        for student in students:
            row = {
                '이름': student.user.name,
                '이메일': student.user.email,
                '연락처': student.user.phone or '',
                '비상연락망': student.emergency_contact or '',
                '주소': student.address or '',
                '지점명': student.branch_name,
                '클래스명': student.class_name or '',
                '시간대': student.time_slot or '',
                '승인상태': '승인완료' if student.is_approved else '승인대기',
                '수강시작일': student.enrollment_date.strftime('%Y-%m-%d') if student.enrollment_date else '',
                '등록일': student.user.created_at.strftime('%Y-%m-%d')
            }
            data.append(row)
        
        # DataFrame 생성 및 엑셀로 변환
        df = pd.DataFrame(data)
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='회원명부')
        output.seek(0)
        
        # 파일명 생성
        branch_name = current_user.managed_branch.name if current_user.role == 'admin' else '전체'
        filename = f"{branch_name}_회원명부_{date.today().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f"회원 명부 다운로드 중 오류가 발생했습니다: {str(e)}", "danger")
        return redirect(url_for('manage_students'))

@app.route('/download_template')
@admin_required
def download_template():
    """기존 회원명단 + 업로드 양식 통합 다운로드"""
    try:
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Font, PatternFill, Alignment
        
        current_user = User.query.get(session['user_id'])
        
        # 🔹 수정: 지점별 정확한 처리
        if current_user.role == 'master':
            # 마스터는 모든 지점 데이터
            classes = Class.query.all()
            students = Student.query.join(User).order_by(User.created_at.desc()).all()
            branch_name = "전체지점"
        else:
            # 관리자는 자신의 지점만
            classes = Class.query.filter_by(branch_id=current_user.branch_id).all()
            students = Student.query.filter_by(branch_id=current_user.branch_id).join(User).order_by(User.created_at.desc()).all()
            branch_name = current_user.managed_branch.name if current_user.managed_branch else f"지점{current_user.branch_id}"
        
        # 새 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "회원명부_통합"
        
        # 🎨 스타일 정의
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        existing_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")  # 기존 회원 (연한 녹색)
        template_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 템플릿 (연한 노랑)
        
        # 📋 헤더 설정
        headers = [
            '번호', '이름', '이메일', '초기비밀번호', '연락처', '비상연락망', 
            '주소', '지점명', '클래스명', '시간대', '수강시작일(YYYY-MM-DD)', '수강기간(개월)', '상태'
        ]
        
        # 헤더 추가
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        current_row = 2
        
        # 🔹 기존 회원 데이터 추가
        for idx, student in enumerate(students, 1):
            row_data = [
                idx,
                student.user.name,
                student.user.email,
                "****",  # 보안상 비밀번호는 숨김
                student.user.phone or "",
                student.emergency_contact or "",
                student.address or "",
                student.branch_name,
                student.class_name or "",
                student.time_slot or "",
                student.start_date.strftime('%Y-%m-%d') if student.start_date else "",
                "",  # 수강기간은 기존 데이터에 없음
                "기존회원"
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.fill = existing_fill  # 기존 회원은 녹색
            
            current_row += 1
        
        # 🔹 구분선 추가
        if students:
            ws.merge_cells(f'A{current_row}:M{current_row}')
            separator_cell = ws[f'A{current_row}']
            separator_cell.value = "⬇️⬇️⬇️ 아래에 새로운 회원 정보를 입력하세요 ⬇️⬇️⬇️"
            separator_cell.font = Font(bold=True, color="FF0000")
            separator_cell.alignment = Alignment(horizontal='center')
            separator_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            current_row += 1
        
        # 🔹 빈 템플릿 양식 추가 (5줄)
        template_start_row = current_row
        for i in range(5):
            sample_row = [
                len(students) + i + 1,  # 번호 이어서
                f"신규회원{i+1}",
                f"new{i+1}@example.com",
                "1234",
                "010-0000-0000",
                "010-0000-0000",
                "주소를 입력하세요",
                branch_name,
                "",  # 드롭다운으로 선택
                "",  # 드롭다운으로 선택
                "2025-09-01",
                "3",
                "신규"
            ]
            
            for col_idx, value in enumerate(sample_row, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=value)
                cell.fill = template_fill  # 템플릿은 노란색
            
            current_row += 1
        
        # 🔽 드롭다운 설정
        class_time_options = []
        time_options = set()
        
        for cls in classes:
            for time_slot in cls.time_slots:
                class_time_options.append(cls.name)
                time_options.add(time_slot.time)
        
        unique_classes = sorted(list(set(class_time_options)))
        unique_times = sorted(list(time_options))
        
        if unique_classes:
            class_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(unique_classes)}"',
                showErrorMessage=True,
                errorTitle="클래스명 오류",
                error="등록된 클래스 중에서 선택해주세요."
            )
            ws.add_data_validation(class_validation)
            class_validation.add(f'I{template_start_row}:I1000')
        
        if unique_times:
            time_validation = DataValidation(
                type="list",
                formula1=f'"{",".join(unique_times)}"',
                showErrorMessage=True,
                errorTitle="시간대 오류",
                error="등록된 시간대 중에서 선택해주세요."
            )
            ws.add_data_validation(time_validation)
            time_validation.add(f'J{template_start_row}:J1000')
        
        # 컬럼 너비 조정
        column_widths = {
            'A': 5, 'B': 12, 'C': 25, 'D': 12, 'E': 15, 'F': 15, 
            'G': 40, 'H': 15, 'I': 15, 'J': 10, 'K': 18, 'L': 12, 'M': 8
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 파일 저장
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"{branch_name}_통합회원명부_{date.today().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        flash(f"통합 명부 생성 중 오류가 발생했습니다: {str(e)}", "danger")
        return redirect(url_for('manage_students'))

# ----------------------------------------------------
# 🔹 관리자 페이지 라우트 (지점별 권한 분리 적용)
# ----------------------------------------------------


@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    try:
        current_user = User.query.get(session['user_id'])
        
        # 날짜 변수들을 먼저 정의
        today = date.today()
        seven_days_later = today + relativedelta(days=7)
        first_day_of_month = today.replace(day=1)
        
        # 🔹 디버깅 정보 (기존과 동일하게 유지)
        print(f"🔍 현재 사용자: {current_user.name} ({current_user.role})")
        print(f"🔍 사용자 branch_id: {current_user.branch_id}")
        
        # 전체 데이터 확인 (디버깅용)
        all_students = Student.query.all()
        print(f"🔍 데이터베이스 전체 학생 수: {len(all_students)}")
        for i, student in enumerate(all_students[:3]):  # 처음 3명만 확인
            print(f"  - {student.user.name}: branch_id={student.branch_id}, branch_name='{student.branch_name}'")
        
        if current_user.role == 'master':
            print("📊 마스터 모드: 전체 데이터 표시")
            # 마스터는 전체 통계
            total_students = Student.query.count()
            total_vehicles = Vehicle.query.count()
            expiring_soon_count = Student.query.filter(
                Student.end_date.isnot(None),
                Student.end_date <= seven_days_later, 
                Student.end_date >= today
            ).count()
            
            # 마스터는 전체 신규 학생
            new_students_this_month = User.query.filter(
                User.role == 'student', 
                User.created_at >= first_day_of_month
            ).count()
            
        else:
            print(f"📊 지점 관리자 모드: branch_id {current_user.branch_id}로 필터링")
            
            # 🔹 개선: 우선 branch_id로 시도하되, 백업 로직도 유지
            students_by_id = Student.query.filter_by(branch_id=current_user.branch_id).all()
            print(f"🔍 branch_id로 찾은 학생: {len(students_by_id)}명")
            
            # 백업 방법: branch_name으로도 확인 (데이터 무결성을 위해)
            students_by_name = []
            if hasattr(current_user, 'managed_branch') and current_user.managed_branch:
                students_by_name = Student.query.filter_by(branch_name=current_user.managed_branch.name).all()
                print(f"🔍 branch_name으로 찾은 학생: {len(students_by_name)}명")
                
                # 🔹 개선: 데이터 불일치 경고
                if len(students_by_id) != len(students_by_name):
                    print(f"⚠️ 데이터 불일치 감지: branch_id({len(students_by_id)}) vs branch_name({len(students_by_name)})")
                    # 더 많은 결과를 가진 방법 선택 (기존 로직 유지)
                    if len(students_by_name) > len(students_by_id):
                        print("🔄 branch_name 방식 사용 (더 많은 데이터)")
                        selected_students = students_by_name
                        filter_method = 'branch_name'
                    else:
                        print("🔄 branch_id 방식 사용")
                        selected_students = students_by_id
                        filter_method = 'branch_id'
                else:
                    print("✅ 데이터 일치: branch_id 방식 사용")
                    selected_students = students_by_id
                    filter_method = 'branch_id'
            else:
                print("🔄 branch_id 방식만 사용 (managed_branch 없음)")
                selected_students = students_by_id
                filter_method = 'branch_id'
            
            # 통계 계산 (선택된 필터링 방법에 따라)
            total_students = len(selected_students)
            total_vehicles = Vehicle.query.filter_by(branch_id=current_user.branch_id).count()
            
            # 만료 예정 학생 계산 (선택된 방법에 따라)
            if filter_method == 'branch_name' and current_user.managed_branch:
                expiring_soon_count = Student.query.filter(
                    Student.branch_name == current_user.managed_branch.name,
                    Student.end_date.isnot(None),
                    Student.end_date <= seven_days_later, 
                    Student.end_date >= today
                ).count()
                
                # 지점별 신규 학생 (branch_name 기준)
                new_students_this_month = User.query.join(Student).filter(
                    User.role == 'student',
                    User.created_at >= first_day_of_month,
                    Student.branch_name == current_user.managed_branch.name
                ).count()
            else:
                expiring_soon_count = Student.query.filter(
                    Student.branch_id == current_user.branch_id,
                    Student.end_date.isnot(None),
                    Student.end_date <= seven_days_later, 
                    Student.end_date >= today
                ).count()
                
                # 지점별 신규 학생 (branch_id 기준)
                new_students_this_month = User.query.join(Student).filter(
                    User.role == 'student',
                    User.created_at >= first_day_of_month,
                    Student.branch_id == current_user.branch_id
                ).count()
        
        print(f"📊 최종 통계 - 학생: {total_students}, 차량: {total_vehicles}, 만료예정: {expiring_soon_count}, 신규: {new_students_this_month}")
        
        stats = {
            'total_students': total_students, 
            'new_students_this_month': new_students_this_month, 
            'expiring_soon_count': expiring_soon_count, 
            'total_vehicles': total_vehicles
        }
        return render_template('admin/dashboard.html', stats=stats)
    except Exception as e:
        flash(f"대시보드 로딩 중 오류가 발생했습니다: {str(e)}", "danger")
        print(f"❌ 대시보드 오류: {e}")
        import traceback
        traceback.print_exc()
        return redirect(url_for('login'))

# 🔹 개선된 manage_students 함수 (기존 디버깅 기능 유지)
# app.py의 manage_students 함수를 다음과 같이 수정하세요

@app.route('/admin/students')
@admin_required
def manage_students():
    try:
        current_user = User.query.get(session['user_id'])
        
        print(f"🔍 학생 관리 - 현재 사용자: {current_user.name} ({current_user.role})")
        print(f"🔍 사용자 branch_id: {current_user.branch_id}")
        
        if current_user.role == 'master':
            print("📋 마스터 모드: 모든 학생 조회 + 통계")
            # 마스터는 모든 학생 조회 가능
            all_students = Student.query.join(User).order_by(User.created_at.desc()).all()
            
            # 🔹 마스터 전용 통계 데이터 생성
            # 지점별 통계
            branch_stats = []
            all_branches = Branch.query.all()
            
            for branch in all_branches:
                branch_students = Student.query.filter_by(branch_id=branch.id).all()
                if branch_students:  # 학생이 있는 지점만 포함
                    approved_count = len([s for s in branch_students if s.status == 'approved'])
                    pending_count = len([s for s in branch_students if s.status == 'pending'])
                    
                    # 만료 예정 학생 계산 (7일 이내)
                    today = date.today()
                    seven_days_later = today + relativedelta(days=7)
                    expiring_count = len([s for s in branch_students 
                                        if s.end_date and s.end_date <= seven_days_later and s.end_date >= today])
                    
                    branch_stats.append({
                        'branch_name': branch.name,
                        'count': len(branch_students),
                        'approved': approved_count,
                        'pending': pending_count,
                        'expiring': expiring_count
                    })
            
            # 클래스명 목록 (필터용)
            class_names = list(set([s.class_name for s in all_students if s.class_name]))
            class_names.sort()
            
            # 🔹 마스터 전용 템플릿 사용
            return render_template('admin/manage_students_master.html', 
                                 students=all_students, 
                                 today=date.today(),
                                 branch_stats=branch_stats,
                                 class_names=class_names)
        else:
            print(f"📋 지점 관리자 모드: 지점별 학생 조회")
            
            # 🔹 기존 일반 관리자 로직 (변경 없음)
            students_by_id = Student.query.join(User).filter(
                Student.branch_id == current_user.branch_id
            ).order_by(User.created_at.desc()).all()
            print(f"🔍 branch_id로 찾은 학생: {len(students_by_id)}명")
            
            # 백업: branch_name으로도 확인
            students_by_name = []
            if hasattr(current_user, 'managed_branch') and current_user.managed_branch:
                students_by_name = Student.query.join(User).filter(
                    Student.branch_name == current_user.managed_branch.name
                ).order_by(User.created_at.desc()).all()
                print(f"🔍 branch_name으로 찾은 학생: {len(students_by_name)}명")
                
                if len(students_by_id) != len(students_by_name):
                    print(f"⚠️ 학생 데이터 불일치: branch_id({len(students_by_id)}) vs branch_name({len(students_by_name)})")
                    if len(students_by_name) > len(students_by_id):
                        print("🔄 branch_name 방식으로 학생 조회")
                        all_students = students_by_name
                    else:
                        print("🔄 branch_id 방식으로 학생 조회")
                        all_students = students_by_id
                else:
                    print("✅ 학생 데이터 일치: branch_id 방식 사용")
                    all_students = students_by_id
            else:
                print("🔄 branch_id 방식만으로 학생 조회")
                all_students = students_by_id
            
            # 최후의 수단: 모든 학생을 조회해서 필터링
            if len(all_students) == 0:
                print("🔄 최후의 수단: 모든 학생을 조회해서 필터링")
                all_db_students = Student.query.join(User).order_by(User.created_at.desc()).all()
                filtered_students = []
                for student in all_db_students:
                    if (student.branch_id == current_user.branch_id or 
                        (hasattr(current_user, 'managed_branch') and 
                         current_user.managed_branch and 
                         student.branch_name == current_user.managed_branch.name)):
                        filtered_students.append(student)
                all_students = filtered_students
                print(f"🔍 필터링 후 학생: {len(all_students)}명")
            
            # 🔹 일반 관리자는 기존 템플릿 사용
            return render_template('admin/manage_students.html', students=all_students, today=date.today())
            
    except Exception as e:
        flash(f"학생 목록 로딩 중 오류가 발생했습니다: {str(e)}", "danger")
        print(f"❌ 학생 관리 오류: {e}")
        import traceback
        traceback.print_exc()
        return redirect(url_for('admin_dashboard'))

# 🔹 추가: 마스터 전용 대시보드 통계 API
@app.route('/api/master/branch_growth/<int:branch_id>')
@master_required
def get_branch_growth(branch_id):
    """마스터 전용: 지점별 회원 증감 통계"""
    try:
        # 최근 6개월 데이터
        today = date.today()
        months_data = []
        
        for i in range(5, -1, -1):  # 6개월 전부터 현재까지
            target_date = today - relativedelta(months=i)
            month_start = target_date.replace(day=1)
            month_end = (month_start + relativedelta(months=1)) - relativedelta(days=1)
            
            # 해당 월에 가입한 학생 수
            if branch_id == 0:  # 전체 지점
                new_students = User.query.join(Student).filter(
                    User.role == 'student',
                    User.created_at >= month_start,
                    User.created_at <= month_end
                ).count()
                
                # 해당 월 말 기준 총 학생 수
                total_students = User.query.join(Student).filter(
                    User.role == 'student',
                    User.created_at <= month_end
                ).count()
            else:  # 특정 지점
                new_students = User.query.join(Student).filter(
                    User.role == 'student',
                    Student.branch_id == branch_id,
                    User.created_at >= month_start,
                    User.created_at <= month_end
                ).count()
                
                total_students = User.query.join(Student).filter(
                    User.role == 'student',
                    Student.branch_id == branch_id,
                    User.created_at <= month_end
                ).count()
            
            months_data.append({
                'month': target_date.strftime('%Y-%m'),
                'month_name': target_date.strftime('%m월'),
                'new_students': new_students,
                'total_students': total_students
            })
        
        return jsonify(months_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/master/weekly_stats')
@master_required
def get_weekly_stats():
    """마스터 전용: 주간 통계"""
    try:
        today = date.today()
        stats = []
        
        for i in range(6, -1, -1):  # 7일 전부터 오늘까지
            target_date = today - relativedelta(days=i)
            
            # 해당 날짜 가입자
            daily_signups = User.query.join(Student).filter(
                User.role == 'student',
                func.date(User.created_at) == target_date
            ).count()
            
            # 해당 날짜 승인된 학생
            daily_approvals = Student.query.join(User).filter(
                Student.status == 'approved',
                func.date(User.created_at) == target_date
            ).count()
            
            stats.append({
                'date': target_date.strftime('%Y-%m-%d'),
                'day_name': target_date.strftime('%a'),
                'signups': daily_signups,
                'approvals': daily_approvals
            })
        
        return jsonify(stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 🔹 마스터 전용 대시보드 라우트 개선
@app.route('/master/dashboard')
@master_required
def master_dashboard():
    """마스터 전용 고급 대시보드"""
    try:
        today = date.today()
        seven_days_later = today + relativedelta(days=7)
        first_day_of_month = today.replace(day=1)
        
        # 전체 통계
        total_students = Student.query.count()
        total_vehicles = Vehicle.query.count()
        total_branches = Branch.query.count()
        total_classes = Class.query.count()
        
        # 이번 달 신규 가입
        new_students_this_month = User.query.filter(
            User.role == 'student', 
            User.created_at >= first_day_of_month
        ).count()
        
        # 만료 예정
        expiring_soon = Student.query.filter(
            Student.end_date.isnot(None),
            Student.end_date <= seven_days_later, 
            Student.end_date >= today
        ).count()
        
        # 승인 대기
        pending_approvals = Student.query.filter_by(status='pending').count()
        
        # 지점별 상세 통계
        branch_details = []
        all_branches = Branch.query.all()
        
        for branch in all_branches:
            branch_students = Student.query.filter_by(branch_id=branch.id).all()
            branch_vehicles = Vehicle.query.filter_by(branch_id=branch.id).count()
            branch_classes = Class.query.filter_by(branch_id=branch.id).count()
            
            # 이번 달 신규 (해당 지점)
            branch_new_this_month = User.query.join(Student).filter(
                User.role == 'student',
                Student.branch_id == branch.id,
                User.created_at >= first_day_of_month
            ).count()
            
            # 지난 달 신규 (비교용)
            last_month_start = (first_day_of_month - relativedelta(months=1))
            last_month_end = first_day_of_month - relativedelta(days=1)
            
            branch_new_last_month = User.query.join(Student).filter(
                User.role == 'student',
                Student.branch_id == branch.id,
                User.created_at >= last_month_start,
                User.created_at <= last_month_end
            ).count()
            
            # 증감률 계산
            if branch_new_last_month > 0:
                growth_rate = ((branch_new_this_month - branch_new_last_month) / branch_new_last_month) * 100
            else:
                growth_rate = 100 if branch_new_this_month > 0 else 0
            
            branch_details.append({
                'name': branch.name,
                'id': branch.id,
                'total_students': len(branch_students),
                'vehicles': branch_vehicles,
                'classes': branch_classes,
                'new_this_month': branch_new_this_month,
                'new_last_month': branch_new_last_month,
                'growth_rate': round(growth_rate, 1),
                'approved': len([s for s in branch_students if s.status == 'approved']),
                'pending': len([s for s in branch_students if s.status == 'pending'])
            })
        
        # 성장률 기준으로 정렬
        branch_details.sort(key=lambda x: x['growth_rate'], reverse=True)
        
        stats = {
            'total_students': total_students,
            'total_vehicles': total_vehicles,
            'total_branches': total_branches,
            'total_classes': total_classes,
            'new_students_this_month': new_students_this_month,
            'expiring_soon': expiring_soon,
            'pending_approvals': pending_approvals,
            'branch_details': branch_details
        }
        
        return render_template('master/dashboard_advanced.html', stats=stats, today=today)
        
    except Exception as e:
        flash(f"마스터 대시보드 로딩 중 오류가 발생했습니다: {str(e)}", "danger")
        print(f"❌ 마스터 대시보드 오류: {e}")
        return redirect(url_for('admin_dashboard'))
    
@app.route('/master/advanced-dashboard')
@master_required
def master_advanced_dashboard():
    """마스터 전용 고급 대시보드"""
    try:
        today = date.today()
        
        # 기본 통계만 계산 (나머지는 API로 로드)
        total_students = Student.query.count()
        total_branches = Branch.query.count()
        total_vehicles = Vehicle.query.count()
        new_students_this_month = User.query.filter(
            User.role == 'student', 
            User.created_at >= today.replace(day=1)
        ).count()
        pending_approvals = Student.query.filter_by(status='pending').count()
        
        stats = {
            'total_students': total_students,
            'total_branches': total_branches, 
            'total_vehicles': total_vehicles,
            'new_students_this_month': new_students_this_month,
            'pending_approvals': pending_approvals
        }
        
        return render_template('master/advanced_dashboard.html', stats=stats, today=today)
        
    except Exception as e:
        flash(f"고급 대시보드 로딩 중 오류가 발생했습니다: {str(e)}", "danger")
        return redirect(url_for('manage_branches'))    

@app.route('/admin/approve_student/<int:student_id>', methods=['POST'])
@admin_required
def approve_student(student_id):
    try:
        current_user = User.query.get(session['user_id'])
        student_to_approve = Student.query.get_or_404(student_id)
        
        # 🔹 통합된 권한 체크 사용
        if not check_user_permission_for_student(current_user, student_to_approve):
            flash("해당 학생을 승인할 권한이 없습니다.", "danger")
            return redirect(url_for('manage_students'))
        
        student_to_approve.status = 'approved'
        db.session.commit()
        flash(f"'{student_to_approve.user.name}' 학생의 가입을 승인했습니다.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"승인 처리 중 오류가 발생했습니다: {str(e)}", "danger")
    return redirect(url_for('manage_students'))

@app.route('/admin/delete_student/<int:student_id>', methods=['POST'])
@admin_required
def delete_student(student_id):
    try:
        current_user = User.query.get(session['user_id'])
        student_to_delete = Student.query.get_or_404(student_id)
        
        # 🔹 통합된 권한 체크 사용
        if not check_user_permission_for_student(current_user, student_to_delete):
            flash("해당 학생을 삭제할 권한이 없습니다.", "danger")
            return redirect(url_for('manage_students'))
        
        user_to_delete = User.query.get(student_to_delete.user_id)
        
        db.session.delete(student_to_delete)
        db.session.delete(user_to_delete)
        db.session.commit()
        flash(f"'{user_to_delete.name}' 학생의 정보가 영구적으로 삭제되었습니다.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"삭제 처리 중 오류가 발생했습니다: {str(e)}", "danger")
    return redirect(url_for('manage_students'))


@app.route('/admin/extend_subscription/<int:student_id>', methods=['POST'])
@admin_required
def extend_subscription(student_id):
    try:
        current_user = User.query.get(session['user_id'])
        student = Student.query.get_or_404(student_id)
        
        # 🔹 통합된 권한 체크 사용
        if not check_user_permission_for_student(current_user, student):
            flash("해당 학생의 기간을 연장할 권한이 없습니다.", "danger")
            return redirect(url_for('manage_students'))
        
        months_to_extend = int(request.form.get('months', 1))
        
        if student.end_date:
            student.end_date += relativedelta(months=months_to_extend)
            student.extension_count += 1
            db.session.commit()
            flash(f"'{student.user.name}' 학생의 수강 기간이 {months_to_extend}개월 연장되었습니다.", "success")
        else:
            flash("수강 시작일이 지정되지 않아 기간을 연장할 수 없습니다.", "danger")
    except Exception as e:
        db.session.rollback()
        flash(f"기간 연장 중 오류가 발생했습니다: {str(e)}", "danger")
    return redirect(url_for('manage_students'))

@app.route('/admin/classes', methods=['GET', 'POST'])
@admin_required
def manage_classes():
    current_user = User.query.get(session['user_id'])
    
    if request.method == 'POST':
        try:
            class_name = request.form.get('class_name')
            time_slots_str = request.form.get('time_slots')
            durations_str = request.form.get('durations')
            branch_id = request.form.get('branch_id')
            
            if not all([class_name, time_slots_str, durations_str, branch_id]):
                flash("지점, 클래스 이름, 시간대, 수강 기간을 모두 입력해주세요.", "danger")
                return redirect(url_for('manage_classes'))
            
            # 권한 체크: 일반 관리자는 자신의 지점에만 클래스 생성 가능
            if current_user.role != 'master' and int(branch_id) != current_user.branch_id:
                flash("해당 지점에 클래스를 생성할 권한이 없습니다.", "danger")
                return redirect(url_for('manage_classes'))
                
            # 지점 존재 확인
            branch = Branch.query.get(branch_id)
            if not branch:
                flash("선택한 지점이 존재하지 않습니다.", "danger")
                return redirect(url_for('manage_classes'))
            
            time_slots_list = [t.strip() for t in time_slots_str.split(',') if t.strip()]
            
            new_class = Class(name=class_name, durations=durations_str, branch_id=int(branch_id))
            db.session.add(new_class)
            db.session.flush()  # ID 생성을 위해 flush
            
            for time_str in time_slots_list:
                new_slot = TimeSlot(time=time_str, class_id=new_class.id)
                db.session.add(new_slot)
                
            db.session.commit()
            flash(f"'{class_name}' 클래스가 성공적으로 생성되었습니다.", "success")
            
        except Exception as e:
            db.session.rollback()
            flash(f"클래스 생성 중 오류가 발생했습니다: {str(e)}", "danger")
            
        return redirect(url_for('manage_classes'))
    
    # GET 요청 처리 - 지점별 필터링 적용
    available_times = []
    for hour in range(6, 23):
        available_times.append(f"{hour:02d}:00")
        available_times.append(f"{hour:02d}:30")

    if current_user.role == 'master':
        # 마스터는 모든 클래스 조회 가능
        all_classes = Class.query.order_by(Class.id.desc()).all()
        all_branches = Branch.query.all()
    else:
        # 일반 관리자는 자신의 지점 클래스만 조회
        all_classes = Class.query.filter_by(branch_id=current_user.branch_id).order_by(Class.id.desc()).all()
        all_branches = [current_user.managed_branch]  # 자신의 지점만

    return render_template('admin/manage_classes.html', 
                         classes=all_classes, 
                         available_times=available_times, 
                         branches=all_branches,
                         current_user=current_user)

# app.py - API 라우트 수정 (기존 코드 교체)

@app.route('/admin/delete_class/<int:class_id>', methods=['POST'])
@admin_required
def delete_class(class_id):
    try:
        current_user = User.query.get(session['user_id'])
        class_to_delete = Class.query.get_or_404(class_id)
        
        # 🔹 권한 체크: 마스터이거나 해당 지점의 클래스만 삭제 가능
        if not check_user_permission_for_class(current_user, class_to_delete):
            flash("해당 클래스를 삭제할 권한이 없습니다.", "danger")
            return redirect(url_for('manage_classes'))
        
        # 해당 클래스에 등록된 학생이 있는지 확인
        students_in_class = Student.query.filter_by(class_name=class_to_delete.name).count()
        if students_in_class > 0:
            flash(f"'{class_to_delete.name}' 클래스에 등록된 학생({students_in_class}명)이 있어 삭제할 수 없습니다.", "danger")
            return redirect(url_for('manage_classes'))
            
        class_name = class_to_delete.name
        
        # 관련된 시간대들도 함께 삭제 (CASCADE)
        TimeSlot.query.filter_by(class_id=class_id).delete()
        
        db.session.delete(class_to_delete)
        db.session.commit()
        flash(f"'{class_name}' 클래스가 삭제되었습니다.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"클래스 삭제 중 오류가 발생했습니다: {str(e)}", "danger")
        
    return redirect(url_for('manage_classes'))

@app.route('/api/classes_by_branch/<int:branch_id>')
@login_required
def get_classes_by_branch(branch_id):
    try:
        current_user = User.query.get(session['user_id'])
        
        print(f"🔍 클래스 조회 요청 - 사용자: {current_user.name if current_user else 'None'}, 지점ID: {branch_id}")
        
        # 🔹 회원가입 중인 학생은 모든 지점 클래스 조회 가능하도록 우선 허용
        if current_user.role == 'student' or not hasattr(current_user, 'role'):
            print("📋 회원가입 모드: 모든 지점 클래스 조회 허용")
            classes = Class.query.filter_by(branch_id=branch_id).all()
        elif current_user.role == 'master':
            print("📋 마스터 모드: 모든 지점 클래스 조회")
            classes = Class.query.filter_by(branch_id=branch_id).all()
        elif current_user.role == 'admin':
            print(f"📋 관리자 모드: 지점 권한 체크 ({current_user.branch_id} vs {branch_id})")
            if current_user.branch_id != branch_id:
                print("❌ 권한 없음")
                return jsonify({'error': '권한이 없습니다.'}), 403
            classes = Class.query.filter_by(branch_id=branch_id).all()
        else:
            print("❌ 알 수 없는 역할")
            return jsonify({'error': '권한이 없습니다.'}), 403
        
        print(f"📊 조회된 클래스 수: {len(classes)}")
        for cls in classes:
            print(f"  - {cls.name} (ID: {cls.id})")
            
        class_list = [{'id': c.id, 'name': c.name} for c in classes]
        return jsonify(class_list)
    except Exception as e:
        print(f"❌ 클래스 조회 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/class_info/<int:class_id>')
@login_required
def get_class_info(class_id):
    try:
        current_user = User.query.get(session['user_id'])
        class_item = Class.query.get_or_404(class_id)
        
        print(f"🔍 클래스 정보 요청 - 클래스ID: {class_id}, 사용자: {current_user.name if current_user else 'None'}")
        
        # 🔹 회원가입 중인 학생은 모든 클래스 정보 조회 가능하도록 우선 허용
        if current_user.role == 'student' or not hasattr(current_user, 'role'):
            print("📋 회원가입 모드: 모든 클래스 정보 조회 허용")
            pass
        elif current_user.role == 'master':
            print("📋 마스터 모드: 모든 클래스 정보 조회")
            pass
        elif current_user.role == 'admin':
            print(f"📋 관리자 모드: 지점 권한 체크 ({current_user.branch_id} vs {class_item.branch_id})")
            if current_user.branch_id != class_item.branch_id:
                print("❌ 권한 없음")
                return jsonify({'error': '권한이 없습니다.'}), 403
        else:
            print("❌ 알 수 없는 역할")
            return jsonify({'error': '권한이 없습니다.'}), 403
            
        time_slots = sorted([slot.time for slot in class_item.time_slots])
        durations = sorted([int(d.strip()) for d in class_item.durations.split(',') if d.strip()]) if class_item.durations else []
        
        print(f"📊 클래스 정보: 시간대 {len(time_slots)}개, 기간 옵션 {len(durations)}개")
        
        return jsonify({'time_slots': time_slots, 'durations': durations})
    except Exception as e:
        print(f"❌ 클래스 정보 조회 오류: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# 🔹 추가: 회원가입 전용 API (로그인 불필요)
@app.route('/api/public/classes_by_branch/<int:branch_id>')
def get_public_classes_by_branch(branch_id):
    """회원가입 시 사용하는 공개 API (로그인 불필요)"""
    try:
        print(f"🔍 공개 클래스 조회 요청 - 지점ID: {branch_id}")
        
        classes = Class.query.filter_by(branch_id=branch_id).all()
        print(f"📊 조회된 클래스 수: {len(classes)}")
        
        class_list = [{'id': c.id, 'name': c.name} for c in classes]
        return jsonify(class_list)
    except Exception as e:
        print(f"❌ 공개 클래스 조회 오류: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/public/class_info/<int:class_id>')
def get_public_class_info(class_id):
    """회원가입 시 사용하는 공개 API (로그인 불필요)"""
    try:
        print(f"🔍 공개 클래스 정보 요청 - 클래스ID: {class_id}")
        
        class_item = Class.query.get_or_404(class_id)
        time_slots = sorted([slot.time for slot in class_item.time_slots])
        durations = sorted([int(d.strip()) for d in class_item.durations.split(',') if d.strip()]) if class_item.durations else []
        
        print(f"📊 클래스 정보: 시간대 {len(time_slots)}개, 기간 옵션 {len(durations)}개")
        
        return jsonify({'time_slots': time_slots, 'durations': durations})
    except Exception as e:
        print(f"❌ 공개 클래스 정보 조회 오류: {str(e)}")
        return jsonify({'error': str(e)}), 500

# app.py에 추가할 API 라우트들 (기존 API 라우트들 아래에 추가하세요)

@app.route('/api/branch-stats')
@admin_required
def get_branch_stats():
    """지점별 학생 분포 통계"""
    try:
        current_user = User.query.get(session['user_id'])
        
        if current_user.role == 'master':
            # 마스터는 모든 지점 통계
            branches = Branch.query.all()
            stats = []
            for branch in branches:
                student_count = Student.query.filter_by(branch_id=branch.id).count()
                if student_count > 0:  # 학생이 있는 지점만 포함
                    stats.append({
                        'name': branch.name,
                        'count': student_count
                    })
        else:
            # 일반 관리자는 자신의 지점만
            if current_user.managed_branch:
                student_count = Student.query.filter_by(branch_id=current_user.branch_id).count()
                stats = [{
                    'name': current_user.managed_branch.name,
                    'count': student_count
                }]
            else:
                stats = []
        
        # Chart.js 형식으로 변환
        labels = [stat['name'] for stat in stats]
        values = [stat['count'] for stat in stats]
        
        return jsonify({
            'labels': labels,
            'values': values
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/monthly-stats')
@admin_required
def get_monthly_stats():
    """월별 신규 가입 통계"""
    try:
        current_user = User.query.get(session['user_id'])
        today = date.today()
        
        months_data = []
        month_labels = []
        
        # 최근 6개월 데이터
        for i in range(5, -1, -1):
            target_date = today - relativedelta(months=i)
            month_start = target_date.replace(day=1)
            month_end = (month_start + relativedelta(months=1)) - relativedelta(days=1)
            
            if current_user.role == 'master':
                # 마스터는 전체 신규 가입
                new_students = User.query.join(Student).filter(
                    User.role == 'student',
                    User.created_at >= month_start,
                    User.created_at <= month_end
                ).count()
            else:
                # 일반 관리자는 자신의 지점만
                new_students = User.query.join(Student).filter(
                    User.role == 'student',
                    Student.branch_id == current_user.branch_id,
                    User.created_at >= month_start,
                    User.created_at <= month_end
                ).count()
            
            months_data.append(new_students)
            month_labels.append(target_date.strftime('%m월'))
        
        return jsonify({
            'months': month_labels,
            'signups': months_data
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/detailed-branch-stats')
@master_required
def get_detailed_branch_stats():
    """마스터 전용: 지점별 상세 통계"""
    try:
        today = date.today()
        first_day_of_month = today.replace(day=1)
        
        branches = Branch.query.all()
        detailed_stats = []
        
        for branch in branches:
            # 지점별 학생 수
            branch_students = Student.query.filter_by(branch_id=branch.id).all()
            total_students = len(branch_students)
            
            if total_students == 0:
                continue  # 학생이 없는 지점은 제외
            
            # 승인 상태별 분류
            approved = len([s for s in branch_students if s.status == 'approved'])
            pending = len([s for s in branch_students if s.status == 'pending'])
            
            # 차량 및 클래스 수
            vehicles = Vehicle.query.filter_by(branch_id=branch.id).count()
            classes = Class.query.filter_by(branch_id=branch.id).count()
            
            # 이번 달 신규 학생
            new_this_month = User.query.join(Student).filter(
                User.role == 'student',
                Student.branch_id == branch.id,
                User.created_at >= first_day_of_month
            ).count()
            
            detailed_stats.append({
                'name': branch.name,
                'total_students': total_students,
                'approved': approved,
                'pending': pending,
                'vehicles': vehicles,
                'classes': classes,
                'new_this_month': new_this_month
            })
        
        # 학생 수 기준으로 정렬
        detailed_stats.sort(key=lambda x: x['total_students'], reverse=True)
        
        return jsonify(detailed_stats)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/class-distribution')
@admin_required
def get_class_distribution():
    """클래스별 학생 분포 통계"""
    try:
        current_user = User.query.get(session['user_id'])
        
        if current_user.role == 'master':
            # 마스터는 모든 클래스
            students = Student.query.filter_by(status='approved').all()
        else:
            # 일반 관리자는 자신의 지점만
            students = Student.query.filter_by(
                branch_id=current_user.branch_id,
                status='approved'
            ).all()
        
        # 클래스별 집계
        class_counts = defaultdict(int)
        for student in students:
            if student.class_name:
                class_counts[student.class_name] += 1
        
        # Chart.js 형식으로 변환
        labels = list(class_counts.keys())
        values = list(class_counts.values())
        
        return jsonify({
            'labels': labels,
            'values': values
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/time-distribution')
@admin_required
def get_time_distribution():
    """시간대별 학생 분포 통계"""
    try:
        current_user = User.query.get(session['user_id'])
        
        if current_user.role == 'master':
            # 마스터는 모든 시간대
            students = Student.query.filter_by(status='approved').all()
        else:
            # 일반 관리자는 자신의 지점만
            students = Student.query.filter_by(
                branch_id=current_user.branch_id,
                status='approved'
            ).all()
        
        # 시간대별 집계
        time_counts = defaultdict(int)
        for student in students:
            if student.time_slot:
                time_counts[student.time_slot] += 1
        
        # 시간순으로 정렬
        sorted_times = sorted(time_counts.items(), key=lambda x: x[0])
        
        return jsonify({
            'labels': [time for time, count in sorted_times],
            'values': [count for time, count in sorted_times]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
# 🔥 마스터 전용 고급 통계 API들 - app.py의 맨 마지막 API 다음에 추가

@app.route('/api/master/yearly-growth-comparison')
@master_required
def get_yearly_growth_comparison():
    """마스터 전용: 연도별 성장 비교 (올해 vs 작년)"""
    try:
        today = date.today()
        current_year = today.year
        last_year = current_year - 1
        
        months_data = []
        
        for month in range(1, 13):  # 1월부터 12월까지
            # 올해 해당 월
            current_year_start = date(current_year, month, 1)
            if month == 12:
                current_year_end = date(current_year + 1, 1, 1) - relativedelta(days=1)
            else:
                current_year_end = date(current_year, month + 1, 1) - relativedelta(days=1)
            
            # 작년 해당 월
            last_year_start = date(last_year, month, 1)
            if month == 12:
                last_year_end = date(last_year + 1, 1, 1) - relativedelta(days=1)
            else:
                last_year_end = date(last_year, month + 1, 1) - relativedelta(days=1)
            
            # 미래 월은 건너뛰기
            if current_year_start > today:
                break
                
            # 올해 신규 가입
            current_signups = User.query.join(Student).filter(
                User.role == 'student',
                User.created_at >= current_year_start,
                User.created_at <= current_year_end
            ).count()
            
            # 작년 신규 가입
            last_year_signups = User.query.join(Student).filter(
                User.role == 'student',
                User.created_at >= last_year_start,
                User.created_at <= last_year_end
            ).count()
            
            months_data.append({
                'month': f'{month}월',
                'current_year': current_signups,
                'last_year': last_year_signups,
                'growth_rate': ((current_signups - last_year_signups) / max(last_year_signups, 1)) * 100
            })
        
        return jsonify({
            'data': months_data,
            'current_year': current_year,
            'last_year': last_year
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/master/branch-class-matrix')
@master_required
def get_branch_class_matrix():
    """마스터 전용: 지점별-클래스별 매트릭스 분석"""
    try:
        # 모든 지점과 클래스 조회
        branches = Branch.query.all()
        all_classes = Class.query.all()
        
        matrix_data = []
        class_names = set()
        
        for branch in branches:
            branch_data = {'branch': branch.name}
            
            # 지점별 클래스 통계
            for class_item in all_classes:
                if class_item.branch_id == branch.id:
                    student_count = Student.query.filter_by(
                        branch_id=branch.id,
                        class_name=class_item.name,
                        status='approved'
                    ).count()
                    
                    branch_data[class_item.name] = student_count
                    class_names.add(class_item.name)
            
            # 지점별 총계
            branch_total = Student.query.filter_by(
                branch_id=branch.id,
                status='approved'
            ).count()
            branch_data['total'] = branch_total
            
            if branch_total > 0:  # 학생이 있는 지점만 포함
                matrix_data.append(branch_data)
        
        return jsonify({
            'matrix': matrix_data,
            'class_names': sorted(list(class_names))
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/master/performance-ranking')
@master_required
def get_performance_ranking():
    """마스터 전용: 지점 성과 랭킹 (다양한 지표)"""
    try:
        today = date.today()
        this_month_start = today.replace(day=1)
        last_month_start = (this_month_start - relativedelta(months=1))
        last_month_end = this_month_start - relativedelta(days=1)
        
        branches = Branch.query.all()
        rankings = []
        
        for branch in branches:
            # 기본 통계
            total_students = Student.query.filter_by(branch_id=branch.id).count()
            approved_students = Student.query.filter_by(
                branch_id=branch.id, 
                status='approved'
            ).count()
            
            # 이번 달 신규
            new_this_month = User.query.join(Student).filter(
                User.role == 'student',
                Student.branch_id == branch.id,
                User.created_at >= this_month_start
            ).count()
            
            # 지난 달 신규
            new_last_month = User.query.join(Student).filter(
                User.role == 'student',
                Student.branch_id == branch.id,
                User.created_at >= last_month_start,
                User.created_at <= last_month_end
            ).count()
            
            # 성장률 계산
            growth_rate = 0
            if new_last_month > 0:
                growth_rate = ((new_this_month - new_last_month) / new_last_month) * 100
            elif new_this_month > 0:
                growth_rate = 100
            
            # 승인율 계산
            approval_rate = 0
            if total_students > 0:
                approval_rate = (approved_students / total_students) * 100
            
            # 차량 활용률
            branch_vehicles = Vehicle.query.filter_by(branch_id=branch.id).count()
            vehicle_utilization = 0
            if branch_vehicles > 0:
                vehicle_utilization = min((approved_students / (branch_vehicles * 15)) * 100, 100)  # 차량당 15명 기준
            
            rankings.append({
                'branch_name': branch.name,
                'total_students': total_students,
                'approved_students': approved_students,
                'new_this_month': new_this_month,
                'new_last_month': new_last_month,
                'growth_rate': round(growth_rate, 1),
                'approval_rate': round(approval_rate, 1),
                'vehicle_count': branch_vehicles,
                'vehicle_utilization': round(vehicle_utilization, 1)
            })
        
        # 성장률 기준 정렬
        rankings.sort(key=lambda x: x['growth_rate'], reverse=True)
        
        return jsonify(rankings)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/master/time-slot-analysis')
@master_required
def get_time_slot_analysis():
    """마스터 전용: 시간대별 수요 분석"""
    try:
        # 시간대별 학생 분포
        students = Student.query.filter_by(status='approved').all()
        
        time_analysis = defaultdict(lambda: {
            'total': 0,
            'branches': defaultdict(int)
        })
        
        for student in students:
            if student.time_slot:
                time_analysis[student.time_slot]['total'] += 1
                if student.branch_id:
                    branch = Branch.query.get(student.branch_id)
                    if branch:
                        time_analysis[student.time_slot]['branches'][branch.name] += 1
        
        # 시간순 정렬
        sorted_times = sorted(time_analysis.items())
        
        result = []
        for time_slot, data in sorted_times:
            result.append({
                'time_slot': time_slot,
                'total_students': data['total'],
                'branch_distribution': dict(data['branches']),
                'utilization_rate': round((data['total'] / max(sum(d['total'] for d in time_analysis.values()), 1)) * 100, 1)
            })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/master/monthly-comparison-detailed')
@master_required
def get_monthly_comparison_detailed():
    """마스터 전용: 상세 월별 비교 (지난달 vs 이번달)"""
    try:
        today = date.today()
        this_month_start = today.replace(day=1)
        last_month_start = (this_month_start - relativedelta(months=1))
        last_month_end = this_month_start - relativedelta(days=1)
        
        branches = Branch.query.all()
        comparison_data = []
        
        for branch in branches:
            # 이번 달 신규 가입
            this_month_new = User.query.join(Student).filter(
                User.role == 'student',
                Student.branch_id == branch.id,
                User.created_at >= this_month_start
            ).count()
            
            # 지난 달 신규 가입
            last_month_new = User.query.join(Student).filter(
                User.role == 'student',
                Student.branch_id == branch.id,
                User.created_at >= last_month_start,
                User.created_at <= last_month_end
            ).count()
            
            # 단순한 승인 통계
            this_month_approved = Student.query.filter(
                Student.branch_id == branch.id,
                Student.status == 'approved'
            ).count()
            
            last_month_approved = 0  # 일단 0으로 설정
            
            # 증감률 계산
            new_growth = 0
            if last_month_new > 0:
                new_growth = ((this_month_new - last_month_new) / last_month_new) * 100
            elif this_month_new > 0:
                new_growth = 100
            
            comparison_data.append({
                'branch_name': branch.name,
                'this_month': {
                    'new': this_month_new,
                    'approved': this_month_approved
                },
                'last_month': {
                    'new': last_month_new,
                    'approved': last_month_approved
                },
                'growth': {
                    'new': round(new_growth, 1),
                    'approved': 0
                }
            })
        
        # 신규 가입 성장률 기준 정렬
        comparison_data.sort(key=lambda x: x['growth']['new'], reverse=True)
        
        return jsonify({
            'comparison': comparison_data,
            'period': {
                'current': this_month_start.strftime('%Y년 %m월'),
                'previous': last_month_start.strftime('%Y년 %m월')
            }
        })
    except Exception as e:
        print(f"❌ 월별 비교 API 오류: {str(e)}")
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/master/class-popularity-trends')
@master_required
def get_class_popularity_trends():
    """마스터 전용: 클래스별 인기도 트렌드 (최근 6개월)"""
    try:
        today = date.today()
        months_data = []
        
        # 최근 6개월
        for i in range(5, -1, -1):
            target_date = today - relativedelta(months=i)
            month_start = target_date.replace(day=1)
            month_end = (month_start + relativedelta(months=1)) - relativedelta(days=1)
            
            # 해당 월 신규 가입자들의 클래스 분포
            new_students = Student.query.join(User).filter(
                User.role == 'student',
                User.created_at >= month_start,
                User.created_at <= month_end
            ).all()
            
            class_distribution = defaultdict(int)
            for student in new_students:
                if student.class_name:
                    class_distribution[student.class_name] += 1
            
            months_data.append({
                'month': target_date.strftime('%Y-%m'),
                'month_name': target_date.strftime('%m월'),
                'classes': dict(class_distribution),
                'total_new': len(new_students)
            })
        
        return jsonify(months_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500    
    
# 🔹 개선된 manage_vehicles 함수
@app.route('/admin/vehicles')
@admin_required
def manage_vehicles():
    try:
        current_user = User.query.get(session['user_id'])
        
        print(f"🔍 차량 관리 - 현재 사용자: {current_user.name} ({current_user.role})")
        
        if current_user.role == 'master':
            # 마스터는 모든 차량과 기사 조회 가능
            all_vehicles = Vehicle.query.order_by(Vehicle.id).all()
            all_drivers = User.query.filter_by(role='driver').all()
            all_branches = Branch.query.all()
            print(f"📋 마스터 모드: 전체 차량 {len(all_vehicles)}대, 전체 기사 {len(all_drivers)}명")
        else:
            # 일반 관리자는 자신의 지점 차량과 기사만 조회
            if not current_user.branch_id:
                flash("관리자의 지점이 설정되지 않았습니다.", "danger")
                return redirect(url_for('admin_dashboard'))
                
            all_vehicles = Vehicle.query.filter_by(branch_id=current_user.branch_id).order_by(Vehicle.id).all()
            
            # 🔹 개선: 확실한 지점별 기사 필터링
            all_drivers = User.query.filter(
                User.role == 'driver',
                User.driver_branch_id == current_user.branch_id
            ).all()
                
            all_branches = [current_user.managed_branch] if current_user.managed_branch else []
            print(f"📋 지점 관리자 모드: 지점 차량 {len(all_vehicles)}대, 지점 기사 {len(all_drivers)}명")
            
        return render_template('admin/manage_vehicles.html', 
                             vehicles=all_vehicles, 
                             drivers=all_drivers, 
                             branches=all_branches,
                             current_user=current_user)
    except Exception as e:
        flash(f"차량 목록 로딩 중 오류가 발생했습니다: {str(e)}", "danger")
        print(f"❌ 차량 관리 오류: {e}")
        return redirect(url_for('admin_dashboard'))
@app.route('/admin/add_driver', methods=['POST'])
@admin_required
def add_driver():
    try:
        current_user = User.query.get(session['user_id'])
        form = request.form
        email = form.get('email')
        name = form.get('name')
        phone = form.get('phone')
        password = form.get('password')
        
        if not all([email, name, phone, password]):
            flash("모든 필드를 입력해주세요.", "danger")
            return redirect(url_for('manage_vehicles'))
            
        if User.query.filter_by(email=email).first():
            flash(f"이미 사용 중인 이메일({email})입니다.", "danger")
            return redirect(url_for('manage_vehicles'))
        
        # 🔹 기사가 소속될 지점 결정 (개선된 버전)
        if current_user.role == 'master':
            branch_id = form.get('branch_id')
            if not branch_id:
                flash("지점을 선택해주세요.", "danger")
                return redirect(url_for('manage_vehicles'))
            
            # 🔹 추가: 선택한 지점이 실제로 존재하는지 확인
            selected_branch = Branch.query.get(branch_id)
            if not selected_branch:
                flash("선택한 지점이 존재하지 않습니다.", "danger")
                return redirect(url_for('manage_vehicles'))
        else:
            # 🔹 추가: 일반 관리자의 지점 설정 확인
            if not current_user.branch_id:
                flash("관리자의 지점이 설정되지 않았습니다. 시스템 관리자에게 문의하세요.", "danger")
                return redirect(url_for('manage_vehicles'))
            branch_id = current_user.branch_id
            
        new_driver = User(
            email=email, 
            name=name, 
            phone=phone, 
            role='driver',
            driver_branch_id=int(branch_id)  # 🔹 추가
        )
        new_driver.set_password(password)
        db.session.add(new_driver)
        db.session.commit()
        
        # 🔹 개선: 성공 메시지에 지점 정보 포함
        branch_name = Branch.query.get(branch_id).name
        flash(f"기사 '{new_driver.name}' 계정이 '{branch_name}' 지점에 생성되었습니다.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"기사 생성 중 오류가 발생했습니다: {str(e)}", "danger")
        
    return redirect(url_for('manage_vehicles'))

@app.route('/admin/add_vehicle', methods=['POST'])
@admin_required
def add_vehicle():
    try:
        current_user = User.query.get(session['user_id'])
        form = request.form
        vehicle_number = form.get('vehicle_number')
        capacity = form.get('capacity')
        
        if not vehicle_number or not capacity:
            flash("차량번호와 정원을 모두 입력해주세요.", "danger")
            return redirect(url_for('manage_vehicles'))
            
        if Vehicle.query.filter_by(vehicle_number=vehicle_number).first():
            flash(f"이미 등록된 차량({vehicle_number})입니다.", "danger")
            return redirect(url_for('manage_vehicles'))
        
        # 🔹 수정: 차량이 소속될 지점 결정
        if current_user.role == 'master':
            branch_id = form.get('branch_id')
            if not branch_id:
                flash("지점을 선택해주세요.", "danger")
                return redirect(url_for('manage_vehicles'))
        else:
            branch_id = current_user.branch_id
            
        new_vehicle = Vehicle(vehicle_number=vehicle_number, capacity=int(capacity), branch_id=int(branch_id))
        db.session.add(new_vehicle)
        db.session.commit()
        flash(f"'{new_vehicle.vehicle_number}' 차량이 등록되었습니다.", "success")
        
    except ValueError:
        flash("정원은 숫자로 입력해주세요.", "danger")
    except Exception as e:
        db.session.rollback()
        flash(f"차량 등록 중 오류가 발생했습니다: {str(e)}", "danger")
        
    return redirect(url_for('manage_vehicles'))

@app.route('/admin/assign_driver/<int:vehicle_id>', methods=['POST'])
@admin_required
def assign_driver(vehicle_id):
    try:
        current_user = User.query.get(session['user_id'])
        vehicle = Vehicle.query.get_or_404(vehicle_id)
        
        # 권한 체크: 마스터이거나 해당 지점의 차량만 기사 배정 가능
        if not check_user_permission_for_vehicle(current_user, vehicle):
            flash("해당 차량에 기사를 배정할 권한이 없습니다.", "danger")
            return redirect(url_for('manage_vehicles'))
        
        driver_id = request.form.get('driver_id')
        
        if driver_id == '0':
            vehicle.driver_id = None
        else:
            # 기사 존재 확인
            driver = User.query.filter_by(id=int(driver_id), role='driver').first()
            if not driver:
                flash("선택한 기사가 존재하지 않습니다.", "danger")
                return redirect(url_for('manage_vehicles'))
            vehicle.driver_id = int(driver_id)
            
        db.session.commit()
        flash(f"'{vehicle.vehicle_number}'의 담당 기사가 변경되었습니다.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"기사 배정 중 오류가 발생했습니다: {str(e)}", "danger")
        
    return redirect(url_for('manage_vehicles'))

@app.route('/admin/delete_vehicle/<int:vehicle_id>', methods=['POST'])
@admin_required
def delete_vehicle(vehicle_id):
    try:
        current_user = User.query.get(session['user_id'])
        vehicle_to_delete = Vehicle.query.get_or_404(vehicle_id)
        
        # 권한 체크: 마스터이거나 해당 지점의 차량만 삭제 가능
        # 🔹 통합된 권한 체크 사용
        if not check_user_permission_for_vehicle(current_user, vehicle_to_delete):
            flash("해당 차량을 삭제할 권한이 없습니다.", "danger")
            return redirect(url_for('manage_vehicles'))
        
        vehicle_number = vehicle_to_delete.vehicle_number
        
        # 배차 기록이 있는지 확인
        dispatch_count = DispatchResult.query.filter_by(vehicle_id=vehicle_id).count()
        if dispatch_count > 0:
            flash(f"'{vehicle_number}' 차량은 배차 기록이 있어 삭제할 수 없습니다. 먼저 관련 배차 기록을 삭제해주세요.", "danger")
            return redirect(url_for('manage_vehicles'))
            
        db.session.delete(vehicle_to_delete)
        db.session.commit()
        flash(f"'{vehicle_number}' 차량이 삭제되었습니다.", "success")
        
    except Exception as e:
        db.session.rollback()
        flash(f"차량 삭제 중 오류가 발생했습니다: {str(e)}", "danger")
        
    return redirect(url_for('manage_vehicles'))

@app.route('/admin/dispatch')
@admin_required
def manage_dispatch():
   try:
       current_user = User.query.get(session['user_id'])
       
       # 지점별 클래스 목록 조회
       if current_user.role == 'master':
           available_classes = Class.query.all()
           dispatch_dates = db.session.query(DispatchResult.dispatch_date).distinct().order_by(DispatchResult.dispatch_date.desc()).all()
           available_drivers = User.query.filter_by(role='driver').all()
           available_vehicles = Vehicle.query.all()
       else:
           available_classes = Class.query.filter_by(branch_id=current_user.branch_id).all()
           branch_vehicle_ids = [v.id for v in Vehicle.query.filter_by(branch_id=current_user.branch_id).all()]
           dispatch_dates = db.session.query(DispatchResult.dispatch_date).filter(
               DispatchResult.vehicle_id.in_(branch_vehicle_ids)
           ).distinct().order_by(DispatchResult.dispatch_date.desc()).all()
           available_drivers = User.query.filter(
               User.role == 'driver',
               User.driver_branch_id == current_user.branch_id
           ).all()
           available_vehicles = Vehicle.query.filter_by(branch_id=current_user.branch_id).all()
           
       return render_template('admin/manage_dispatch.html', 
                              dispatch_dates=dispatch_dates,
                              available_classes=available_classes,
                              available_drivers=available_drivers,
                              available_vehicles=available_vehicles,
                              current_user=current_user,
                              today_str=date.today().strftime('%Y-%m-%d'))
   except Exception as e:
       flash(f"배차 관리 페이지 로딩 중 오류가 발생했습니다: {str(e)}", "danger")
       return redirect(url_for('admin_dashboard'))

# ===== 새로운 배차 시스템 API =====

# app.py - 수정된 배차 API 엔드포인트

@app.route('/api/dispatch/regular', methods=['POST'])
@admin_required
def create_regular_dispatch():
    """정규 배차 생성 - 실제 DispatchResult 데이터 저장"""
    try:
        current_user = User.query.get(session['user_id'])
        data = request.get_json()
        
        class_name = data.get('class_name')
        dispatch_date_str = data.get('dispatch_date')
        auto_optimize = data.get('auto_optimize', True)
        auto_assign = data.get('auto_assign', True)
        
        # 날짜 파싱
        try:
            dispatch_date = datetime.strptime(dispatch_date_str, '%Y-%m-%d').date()
        except:
            dispatch_date = date.today()
        
        print(f"🚐 정규배차 생성 시작: {class_name}, {dispatch_date}")
        
        # 해당 날짜에 이미 배차가 있는지 확인
        existing_dispatch = DispatchResult.query.filter_by(
            dispatch_date=dispatch_date
        ).first()
        
        if existing_dispatch:
            return jsonify({
                'success': False, 
                'error': f'{dispatch_date} 날짜에 이미 배차가 존재합니다.'
            })
        
        # 권한별 학생 조회
        if current_user.role == 'master':
            students = Student.query.filter(
                Student.class_name == class_name,
                Student.status == 'approved'
            ).all()
        else:
            students = Student.query.filter(
                Student.class_name == class_name,
                Student.branch_id == current_user.branch_id,
                Student.status == 'approved'
            ).all()
        
        if not students:
            return jsonify({
                'success': False, 
                'error': f'{class_name} 클래스에 승인된 학생이 없습니다.'
            })
        
        # 가용 차량 조회
        if current_user.role == 'master':
            available_vehicles = Vehicle.query.filter(
                Vehicle.driver_id.isnot(None)
            ).all()
        else:
            available_vehicles = Vehicle.query.filter(
                Vehicle.branch_id == current_user.branch_id,
                Vehicle.driver_id.isnot(None)
            ).all()
        
        if not available_vehicles:
            return jsonify({
                'success': False, 
                'error': '기사가 배정된 가용 차량이 없습니다.'
            })
        
        print(f"📊 대상 학생: {len(students)}명, 가용 차량: {len(available_vehicles)}대")
        
        # 실제 배차 데이터 생성 및 저장
        created_count = 0
        
        for i, student in enumerate(students):
            # 차량 순환 배정
            vehicle = available_vehicles[i % len(available_vehicles)]
            
            try:
                # DispatchResult에 실제 데이터 저장
                new_dispatch = DispatchResult(
                    dispatch_date=dispatch_date,
                    student_id=student.id,
                    vehicle_id=vehicle.id,
                    stop_order=i + 1,
                    status='assigned'  # 배정됨 상태
                )
                
                db.session.add(new_dispatch)
                created_count += 1
                
                print(f"  ✅ {student.user.name} → {vehicle.license_plate}")
                
            except Exception as e:
                print(f"  ❌ 배차 생성 실패: {e}")
                continue
        
        # 데이터베이스 커밋
        try:
            db.session.commit()
            print(f"🎉 총 {created_count}건의 배차가 생성되었습니다!")
            
            return jsonify({
                'success': True,
                'message': f'{class_name} 클래스 정규배차가 생성되었습니다.',
                'created_count': created_count,
                'student_count': len(students),
                'vehicles_used': len(set(available_vehicles[:len(students)])),
                'dispatch_date': dispatch_date_str
            })
            
        except Exception as e:
            db.session.rollback()
            print(f"❌ 데이터베이스 저장 실패: {e}")
            return jsonify({
                'success': False,
                'error': f'배차 저장 중 오류가 발생했습니다: {str(e)}'
            })
        
    except Exception as e:
        print(f"❌ 배차 생성 전체 오류: {e}")
        return jsonify({
            'success': False,
            'error': f'배차 생성 중 오류가 발생했습니다: {str(e)}'
        })


@app.route('/api/dispatch/list', methods=['GET'])
@admin_required
def get_dispatch_list():
    """날짜별 배차 목록 조회 - 수정된 버전"""
    try:
        current_user = User.query.get(session['user_id'])
        
        # 날짜 파라미터 가져오기
        date_param = request.args.get('date')
        if date_param:
            try:
                target_date = datetime.strptime(date_param, '%Y-%m-%d').date()
            except:
                target_date = date.today()
        else:
            target_date = date.today()
        
        print(f"📅 배차 목록 조회: {target_date}")
        
        # 권한별 배차 조회
        if current_user.role == 'master':
            dispatches = DispatchResult.query.filter_by(
                dispatch_date=target_date
            ).all()
        else:
            # 해당 지점의 차량들만 조회
            branch_vehicle_ids = [v.id for v in Vehicle.query.filter_by(
                branch_id=current_user.branch_id
            ).all()]
            
            dispatches = DispatchResult.query.filter(
                DispatchResult.dispatch_date == target_date,
                DispatchResult.vehicle_id.in_(branch_vehicle_ids)
            ).all()
        
        print(f"📋 조회된 배차 수: {len(dispatches)}")
        
        # 응답 데이터 구성
        dispatch_list = []
        for dispatch in dispatches:
            # 안전한 데이터 추출
            student_name = '알 수 없음'
            student_class = '미분류'
            vehicle_name = '미배정'
            driver_name = '미배정'
            
            try:
                if dispatch.student and dispatch.student.user:
                    student_name = dispatch.student.user.name
                if dispatch.student and dispatch.student.class_name:
                    student_class = dispatch.student.class_name
                if dispatch.vehicle:
                    vehicle_name = dispatch.vehicle.license_plate or dispatch.vehicle.vehicle_number
                    if dispatch.vehicle.driver_id:
                        driver = User.query.get(dispatch.vehicle.driver_id)
                        if driver:
                            driver_name = driver.name
            except Exception as e:
                print(f"⚠️ 데이터 추출 오류: {e}")
                continue
            
            dispatch_data = {
                'id': dispatch.id,
                'student_name': student_name,
                'class_name': student_class,
                'vehicle_name': vehicle_name,
                'driver_name': driver_name,
                'stop_order': dispatch.stop_order or 0,
                'status': getattr(dispatch, 'status', 'assigned'),
                'dispatch_date': target_date.strftime('%Y-%m-%d')
            }
            
            dispatch_list.append(dispatch_data)
        
        # 정렬 (stop_order 기준)
        dispatch_list.sort(key=lambda x: x['stop_order'])
        
        print(f"✅ 반환할 배차 데이터: {len(dispatch_list)}건")
        
        return jsonify({
            'success': True,
            'dispatches': dispatch_list,
            'total_count': len(dispatch_list),
            'date': target_date.strftime('%Y-%m-%d')
        })
        
    except Exception as e:
        print(f"❌ 배차 목록 조회 오류: {e}")
        return jsonify({
            'success': False,
            'error': f'배차 목록 조회 중 오류가 발생했습니다: {str(e)}',
            'dispatches': []
        })


# 추가: 배차 상태 업데이트 API
@app.route('/api/dispatch/update-status', methods=['POST'])
@admin_required
def update_dispatch_status():
    """배차 상태 업데이트"""
    try:
        data = request.get_json()
        dispatch_id = data.get('dispatch_id')
        new_status = data.get('status')
        
        dispatch = DispatchResult.query.get_or_404(dispatch_id)
        
        # status 필드가 있다면 업데이트
        if hasattr(dispatch, 'status'):
            dispatch.status = new_status
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'배차 상태가 {new_status}로 변경되었습니다.'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'status 필드가 존재하지 않습니다.'
            })
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'상태 업데이트 실패: {str(e)}'
        })
    
@app.route('/api/dispatch/special', methods=['POST'])
@admin_required
def create_special_dispatch():
   try:
       current_user = User.query.get(session['user_id'])
       data = request.get_json()
       
       special_type = data.get('type')
       student_ids = data.get('student_ids', [])
       reason = data.get('reason')
       priority = data.get('priority', 'normal')
       dispatch_date = data.get('date')
       dispatch_time = data.get('time')
       
       if not student_ids:
           return jsonify({'success': False, 'error': '대상 학생을 선택해주세요.'})
       
       if not reason:
           return jsonify({'success': False, 'error': '특별 배차 사유를 입력해주세요.'})
       
       # 특별배차 기록 저장 (임시로 세션에 저장)
       if 'special_dispatches' not in session:
           session['special_dispatches'] = []
       
       special_dispatch = {
           'id': len(session['special_dispatches']) + 1,
           'type': special_type,
           'student_ids': student_ids,
           'reason': reason,
           'priority': priority,
           'date': dispatch_date,
           'time': dispatch_time,
           'status': 'pending',
           'created_by': current_user.name,
           'created_at': datetime.now().isoformat()
       }
       
       session['special_dispatches'].append(special_dispatch)
       session.modified = True
       
       return jsonify({'success': True, 'message': '특별 배차가 신청되었습니다.'})
       
   except Exception as e:
       return jsonify({'success': False, 'error': str(e)})

   
@app.route('/api/students/by-class')
@admin_required
def get_students_by_class():
   try:
       current_user = User.query.get(session['user_id'])
       class_name = request.args.get('class_name')
       
       if current_user.role == 'master':
           students = Student.query.filter(
               Student.class_name == class_name,
               Student.status == 'approved'
           ).all()
       else:
           students = Student.query.filter(
               Student.class_name == class_name,
               Student.branch_id == current_user.branch_id,
               Student.status == 'approved'
           ).all()
       
       student_data = []
       for s in students:
           student_data.append({
               'id': s.id,
               'name': s.user.name,
               'phone': s.user.phone,
               'address': s.address,
               'class_name': s.class_name,
               'time_slot': s.time_slot
           })
           
       return jsonify({'success': True, 'students': student_data})
       
   except Exception as e:
       return jsonify({'success': False, 'error': str(e)})

# 권한 체크 함수
def check_dispatch_permission(user, dispatch_id=None):
   """배차 관련 권한 체크"""
   if user.role == 'master':
       return True
   elif user.role == 'branch_admin':
       return user.branch_id is not None
   elif user.role == 'driver':
       return user.vehicle is not None
   return False

@app.route('/api/dispatch/history')
@admin_required
def get_dispatch_history():
    """배차 이력 조회 API - 개선 버전"""
    try:
        current_user = User.query.get(session['user_id'])
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        
        print(f"📊 배차 이력 API 호출: {from_date} ~ {to_date}")  # 디버그 로그
        
        # 날짜 파싱
        if from_date:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        if to_date:
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
        else:
            to_date = date.today()
            
        # 권한별 배차 이력 조회
        if current_user.role == 'master':
            query = DispatchResult.query
        else:
            branch_vehicle_ids = [v.id for v in Vehicle.query.filter_by(branch_id=current_user.branch_id).all()]
            if not branch_vehicle_ids:
                print("⚠️ 지점에 차량이 없음")
                return jsonify({'success': True, 'history': [], 'total_records': 0})
            query = DispatchResult.query.filter(DispatchResult.vehicle_id.in_(branch_vehicle_ids))
        
        # 날짜 필터 적용
        if from_date:
            query = query.filter(DispatchResult.dispatch_date >= from_date)
        if to_date:
            query = query.filter(DispatchResult.dispatch_date <= to_date)
            
        # 배차 결과 조회
        dispatches = query.order_by(DispatchResult.dispatch_date.desc()).all()
        print(f"📋 조회된 배차 수: {len(dispatches)}")  # 디버그 로그
        
        if not dispatches:
            print("📝 배차 데이터가 없음 - 빈 결과 반환")
            return jsonify({'success': True, 'history': [], 'total_records': 0})
        
        # 날짜별 그룹화
        grouped_dispatches = defaultdict(list)
        for dispatch in dispatches:
            grouped_dispatches[dispatch.dispatch_date].append(dispatch)
        
        # 응답 데이터 구성
        history_data = []
        for dispatch_date, date_dispatches in grouped_dispatches.items():
            try:
                # 안전한 클래스명 추출
                class_stats = defaultdict(int)
                total_students = len(date_dispatches)
                vehicles_used = len(set(d.vehicle_id for d in date_dispatches if d.vehicle_id))
                
                for dispatch in date_dispatches:
                    try:
                        # 여러 방법으로 클래스명 시도
                        class_name = None
                        if hasattr(dispatch, 'class_name') and dispatch.class_name:
                            class_name = dispatch.class_name
                        elif hasattr(dispatch, 'student') and dispatch.student:
                            if hasattr(dispatch.student, 'class_name'):
                                class_name = dispatch.student.class_name
                            elif hasattr(dispatch.student, 'user') and hasattr(dispatch.student.user, 'class_name'):
                                class_name = dispatch.student.user.class_name
                        
                        if class_name:
                            class_stats[class_name] += 1
                        else:
                            class_stats['클래스 미지정'] += 1
                            
                    except Exception as e:
                        print(f"⚠️ 클래스명 추출 오류: {e}")
                        class_stats['오류'] += 1
                
                history_data.append({
                    'date': dispatch_date.strftime('%Y-%m-%d'),
                    'date_formatted': dispatch_date.strftime('%m월 %d일'),
                    'total_students': total_students,
                    'vehicles_used': vehicles_used,
                    'classes': dict(class_stats),
                    'class_summary': ', '.join([f"{cls}({cnt}명)" for cls, cnt in class_stats.items()]) or '배차 정보 없음'
                })
                
            except Exception as e:
                print(f"⚠️ 날짜별 데이터 처리 오류: {e}")
                continue
        
        print(f"✅ 최종 반환 데이터 수: {len(history_data)}")
        
        return jsonify({
            'success': True,
            'history': history_data,
            'total_records': len(history_data)
        })
        
    except Exception as e:
        print(f"❌ 배차 이력 API 오류: {e}")
        return jsonify({'success': False, 'error': str(e)})
    
@app.route('/admin/dispatch/<a_date>')
@login_required
def view_dispatch_by_date(a_date):
   try:
       current_user = User.query.get(session['user_id'])
       target_date = datetime.strptime(a_date, '%Y-%m-%d').date()
       
       if current_user.role == 'master':
           results = DispatchResult.query.filter_by(dispatch_date=target_date).all()
       else:
           branch_vehicle_ids = [v.id for v in Vehicle.query.filter_by(branch_id=current_user.branch_id).all()]
           results = DispatchResult.query.filter(
               DispatchResult.dispatch_date == target_date,
               DispatchResult.vehicle_id.in_(branch_vehicle_ids)
           ).all()
           
       dispatch_by_vehicle = defaultdict(list)
       
       for result in results:
           dispatch_by_vehicle[result.vehicle].append(result)
           
       if not dispatch_by_vehicle:
           return f"<h1>{a_date} 배차 결과</h1><p>해당 날짜의 배차 기록이 없습니다.</p><a href='/admin/dispatch'>돌아가기</a>"
           
       html = f"<h1>{a_date} 배차 결과</h1><a href='/admin/dispatch'>돌아가기</a>"
       for vehicle, students in dispatch_by_vehicle.items():
           html += f"<h2>{vehicle.vehicle_number} ({len(students)}명)</h2><ol>"
           for s in students:
               html += f"<li>{s.student.user.name} ({s.student.address})</li>"
           html += "</ol>"
       return html
       
   except ValueError:
       return "<h1>오류</h1><p>잘못된 날짜 형식입니다.</p><a href='/admin/dispatch'>돌아가기</a>"
   except Exception as e:
       return f"<h1>오류</h1><p>배차 정보 조회 중 오류가 발생했습니다: {str(e)}</p><a href='/admin/dispatch'>돌아가기</a>"

@app.route('/api/test/create-sample-data')
@admin_required
def create_sample_dispatch_data():
    """샘플 배차 데이터 생성"""
    try:
        current_user = User.query.get(session['user_id'])
        
        # 기존 학생과 차량 조회
        students = Student.query.filter_by(status='approved').limit(5).all()
        vehicles = Vehicle.query.limit(2).all()
        
        if not students or not vehicles:
            return jsonify({'success': False, 'error': '학생이나 차량 데이터가 없습니다.'})
        
        # 최근 3일간 샘플 배차 데이터 생성
        from datetime import timedelta
        today = date.today()
        
        for days_ago in range(3):
            target_date = today - timedelta(days=days_ago)
            
            for i, student in enumerate(students[:3]):  # 3명만
                vehicle = vehicles[i % len(vehicles)]
                
                # 기존 배차 데이터 확인
                existing = DispatchResult.query.filter_by(
                    dispatch_date=target_date,
                    student_id=student.id
                ).first()
                
                if not existing:
                    dispatch = DispatchResult(
                        dispatch_date=target_date,
                        student_id=student.id,
                        vehicle_id=vehicle.id,
                        stop_order=i + 1,
                        class_name=student.class_name,
                        student_count=1,
                        created_at=datetime.now()
                    )
                    db.session.add(dispatch)
        
        db.session.commit()
        
        return jsonify({
            'success': True, 
            'message': f'샘플 배차 데이터가 생성되었습니다. 총 {len(students) * 3}개 레코드'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/dispatch/create_today', methods=['POST'])
@admin_required
def create_dispatch_for_today():
   try:
       current_user = User.query.get(session['user_id'])
       today = date.today()
       
       if current_user.role == 'master':
           existing_dispatch = DispatchResult.query.filter_by(dispatch_date=today).first()
       else:
           branch_vehicle_ids = [v.id for v in Vehicle.query.filter_by(branch_id=current_user.branch_id).all()]
           existing_dispatch = DispatchResult.query.filter(
               DispatchResult.dispatch_date == today,
               DispatchResult.vehicle_id.in_(branch_vehicle_ids)
           ).first()
           
       if existing_dispatch:
           flash(f"오늘({today.strftime('%Y-%m-%d')})의 배차 정보는 이미 생성되었습니다.", "warning")
           return redirect(url_for('manage_dispatch'))
       
       if current_user.role == 'master':
           available_vehicles = Vehicle.query.filter(Vehicle.driver_id.isnot(None)).all()
           all_classes = Class.query.all()
       else:
           available_vehicles = Vehicle.query.filter(
               Vehicle.driver_id.isnot(None),
               Vehicle.branch_id == current_user.branch_id
           ).all()
           all_classes = Class.query.filter_by(branch_id=current_user.branch_id).all()
           
       if not available_vehicles:
           flash("운행 가능한 차량(기사가 배정된)이 없습니다.", "danger")
           return redirect(url_for('manage_dispatch'))
           
       total_dispatched_count = 0
       
       for class_item in all_classes:
           for time_slot in class_item.time_slots:
               absent_student_ids = [a.student_id for a in Absence.query.filter_by(absence_date=today).all()]
               
               if current_user.role == 'master':
                   students_to_dispatch = Student.query.filter(
                       Student.status == 'approved', 
                       Student.class_name == class_item.name,
                       Student.time_slot == time_slot.time,
                       Student.id.notin_(absent_student_ids)
                   ).all()
               else:
                   students_to_dispatch = Student.query.filter(
                       Student.status == 'approved', 
                       Student.class_name == class_item.name,
                       Student.time_slot == time_slot.time,
                       Student.branch_id == current_user.branch_id,
                       Student.id.notin_(absent_student_ids)
                   ).all()
               
               if not students_to_dispatch:
                   continue
               
               student_idx = 0
               for vehicle in available_vehicles:
                   for seat_num in range(vehicle.capacity):
                       if student_idx < len(students_to_dispatch):
                           student = students_to_dispatch[student_idx]
                           new_dispatch = DispatchResult(
                               dispatch_date=today,
                               student_id=student.id,
                               vehicle_id=vehicle.id,
                               stop_order=seat_num + 1
                           )
                           db.session.add(new_dispatch)
                           total_dispatched_count += 1
                           student_idx += 1
                       else:
                           break
                   if student_idx >= len(students_to_dispatch):
                       break
       
       if total_dispatched_count > 0:
           db.session.commit()
           flash(f"오늘의 전체 배차가 완료되었습니다. (총 {total_dispatched_count}건)", "success")
       else:
           flash("오늘 배차할 대상 학생이 없습니다.", "info")
           
   except Exception as e:
       db.session.rollback()
       flash(f"배차 생성 중 오류가 발생했습니다: {str(e)}", "danger")
       
   return redirect(url_for('manage_dispatch'))

@app.route('/admin/dispatch/delete/<a_date>', methods=['POST'])
@admin_required
def delete_dispatch_by_date(a_date):
   try:
       current_user = User.query.get(session['user_id'])
       target_date = datetime.strptime(a_date, '%Y-%m-%d').date()
       
       if current_user.role == 'master':
           deleted_count = DispatchResult.query.filter_by(dispatch_date=target_date).count()
           DispatchResult.query.filter_by(dispatch_date=target_date).delete()
       else:
           branch_vehicle_ids = [v.id for v in Vehicle.query.filter_by(branch_id=current_user.branch_id).all()]
           deleted_count = DispatchResult.query.filter(
               DispatchResult.dispatch_date == target_date,
               DispatchResult.vehicle_id.in_(branch_vehicle_ids)
           ).count()
           DispatchResult.query.filter(
               DispatchResult.dispatch_date == target_date,
               DispatchResult.vehicle_id.in_(branch_vehicle_ids)
           ).delete()
           
       db.session.commit()
       flash(f"{a_date}의 배차 정보 {deleted_count}건이 삭제되었습니다.", "success")
       
   except ValueError:
       flash("잘못된 날짜 형식입니다.", "danger")
   except Exception as e:
       db.session.rollback()
       flash(f"배차 삭제 중 오류가 발생했습니다: {str(e)}", "danger")
       
   return redirect(url_for('manage_dispatch'))


@app.route('/driver/view')
@login_required
def driver_view_route():
    try:
        driver_user = User.query.get(session['user_id'])
        
        if driver_user.role != 'driver':
            flash("기사 계정으로만 접근할 수 있습니다.", "danger")
            return redirect(url_for('login'))
            
        if not driver_user.vehicle:
            return render_template('driver/view_route.html', 
                                  students_data=json.dumps([]),
                                  route_info=None, driver=driver_user)
        
        today = date.today()
        vehicle = driver_user.vehicle
        
        todays_route = DispatchResult.query.filter_by(
            dispatch_date=today,
            vehicle_id=vehicle.id
        ).order_by(DispatchResult.stop_order).all()
        
        # 🔥 실제 학생 데이터 변환
        students_data = []
        for i, dispatch in enumerate(todays_route):
            students_data.append({
                'id': dispatch.student.id,
                'name': dispatch.student.user.name,
                'address': dispatch.student.address or '주소 미등록',
                'phone': dispatch.student.user.phone or '연락처 미등록',
                'status': 'pending',
                'estimatedTime': f'{(i+1)*7}분',
                'distance': f'{(i+1)*1.8:.1f}km'
            })
        
        return render_template('driver/view_route.html', 
                              students_data=json.dumps(students_data, ensure_ascii=False),
                              route_info=todays_route, 
                              driver=driver_user, 
                              vehicle=vehicle, 
                              today_str=today.strftime('%Y-%m-%d'))
                              
    except Exception as e:
        flash(f"운행 정보 조회 중 오류가 발생했습니다: {str(e)}", "danger")
        return redirect(url_for('login'))

# 🔹 app.py의 에러 핸들러 수정
@app.errorhandler(404)
def not_found_error(error):
    return "<h1>404 - 페이지를 찾을 수 없습니다</h1>", 404

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return "<h1>500 - 서버 내부 오류</h1>", 500

# 애플리케이션 초기화
with app.app_context():
    try:
        db.create_all()
        setup_initial_accounts()
    except Exception as e:
        print(f"애플리케이션 초기화 오류: {e}")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)